"""
Enrich the 4 KPV gateway workbooks:
1. Add data validation dropdowns to picklist cells
2. Add "Village Reference Rates" comparison block to Assumptions sheets (per-unit + per-ha where derivable)
3. Refresh the Last calibrated stamp

Source for village benchmarks: 2027 Master Project Budgets for Papamoa (114u), Rototuna (131u),
Waihi (96u). KLE is structured per-unit-type without a stage summary rollup; noted but not extracted.
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
import datetime as dt
from pathlib import Path

ROOT = Path("/home/kyle/KPV-Consulting/KPV Feasibilities")
DST_RV1 = ROOT / "KPV Retirement Gateway 1.xlsx"
DST_RV2 = ROOT / "KPV Retirement Gateway 2.xlsx"
DST_RES1 = ROOT / "KPV Residential Gateway 1.xlsx"
DST_RES2 = ROOT / "KPV Residential Gateway 2.xlsx"

YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
LIGHT_GREEN = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
HEADER_BLUE = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
WHITE_BOLD = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
ITALIC_GREY = Font(italic=True, color="595959", size=9)
TODAY = dt.date.today().isoformat()

# Per-unit benchmarks from 2027 Master Project Budgets (all incl GST, NZD)
VILLAGE_RATES = {
    # field: (Papamoa, Rototuna, Waihi)
    "units":                      (114,        131,        96),
    "land_area_ha":               (5.7,        5.5,        4.8),    # approx — verify per project
    "density_units_per_ha":       (20.0,       23.8,       20.0),   # derived
    "construction_total":         (47_233_526, 62_385_255, 50_385_833),
    "construction_per_unit":      (414_329,    476_223,    524_852),
    "landscaping_per_unit":       (18_620,     18_072,     20_017),
    "council_fees_per_unit":      (25_449,     22_087,     26_782),
    "infrastructure_per_unit":    (100_740,    42_869,     108_212),
    "other_site_works_per_unit":  (7_852,      9_463,      11_993),
    "prof_fees_prelim_per_unit":  (16_858,     20_147,     None),    # Waihi not extracted yet
    "clubhouse_total":            (5_297_195,  5_845_866,  None),
    "total_dev_cost":             (71_855_792, 82_986_633, 83_676_529),
    "total_dev_cost_per_unit":    (630_314,    633_486,    871_630),
    "total_sales_value":          (103_548_750,122_820_813,None),
    "total_sales_per_unit":       (908_322,    937_563,    None),
    "gross_margin_pct":           (0.306,      0.324,      None),
}

# Picklist dropdown definitions: (sheet_name, cell_range, options_csv, prompt_title, prompt_message)
RV_LAND_SCENARIOS = '"Ownership,Deferred,Lease,Vendor Finance"'
RES_COUNCILS = '"TCC,WBOPDC,Matamata-Piako DC,SDC,CODC,Hamilton CC,Auckland Council,Other"'
RES_YESNO = '"Yes,No"'
RES_TITLE_TYPES = '"Fee Simple,Cross Lease,Unit Title"'
RES_UNIT_TYPES = '"Stand Alone,Townhouse,Apartment,Duplex"'
RES_TOPOGRAPHY = '"Flat (0-5),Gentle (5-10),Moderate (10-20),Steep (>20)"'
RES_TENURE = '"Freehold,Leasehold"'

def add_dropdown(ws, cell, options_csv, title, msg):
    dv = DataValidation(type="list", formula1=options_csv, allow_blank=True)
    dv.errorTitle = title
    dv.error = msg
    dv.promptTitle = title
    dv.prompt = msg
    ws.add_data_validation(dv)
    dv.add(cell)

def add_village_rates_block(ws, start_row, asset_class):
    """Add Village Reference Rates block. asset_class = 'rv' or 'residential'."""
    r = start_row

    # Section header
    cell = ws.cell(row=r, column=1, value="VILLAGE REFERENCE RATES (KPV portfolio benchmarks, incl GST)")
    cell.font = WHITE_BOLD
    cell.fill = HEADER_BLUE
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    ws.cell(row=r, column=1, value="Source: 2027 Master Project Budgets (Papamoa, Rototuna, Waihi). KLE omitted (per-unit-type structure, no consolidated rollup).").font = ITALIC_GREY
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    ws.cell(row=r, column=1, value="These are reference points only — your project will diverge. Use to sanity-check the assumptions above.").font = ITALIC_GREY
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 2

    # Table header
    headers = ["Metric", "Papamoa", "Rototuna", "Waihi", "KPV avg", "KPV range"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=r, column=i, value=h)
        c.font = WHITE_BOLD
        c.fill = HEADER_BLUE
    r += 1

    # Metrics to show — ordered
    rows = [
        ("Units",                            "units",                    "#,##0"),
        ("Land area (ha, approx)",           "land_area_ha",             "0.0"),
        ("Density (units/ha)",               "density_units_per_ha",     "0.0"),
        ("Construction total ($)",           "construction_total",       "#,##0"),
        ("Construction $/unit",              "construction_per_unit",    "$#,##0"),
        ("Landscaping $/unit",               "landscaping_per_unit",     "$#,##0"),
        ("Council Dev fees $/unit",          "council_fees_per_unit",    "$#,##0"),
        ("Infrastructure $/unit",            "infrastructure_per_unit",  "$#,##0"),
        ("Other site works $/unit",          "other_site_works_per_unit","$#,##0"),
        ("Professional fees + Prelim $/unit","prof_fees_prelim_per_unit","$#,##0"),
        ("Clubhouse / facility total ($)",   "clubhouse_total",          "$#,##0"),
        ("Total Dev Cost ($)",               "total_dev_cost",           "$#,##0"),
        ("Total Dev Cost $/unit",            "total_dev_cost_per_unit",  "$#,##0"),
        ("Total Sales Value ($)",            "total_sales_value",        "$#,##0"),
        ("Sales $/unit",                     "total_sales_per_unit",     "$#,##0"),
        ("Gross margin %",                   "gross_margin_pct",         "0.0%"),
    ]

    for label, key, fmt in rows:
        ws.cell(row=r, column=1, value=label)
        p, ro, w = VILLAGE_RATES[key]
        for i, v in enumerate([p, ro, w], start=2):
            c = ws.cell(row=r, column=i, value=v if v is not None else "")
            if v is not None:
                c.number_format = fmt
        # KPV avg
        non_null = [x for x in (p, ro, w) if x is not None]
        if non_null:
            avg = sum(non_null) / len(non_null)
            c = ws.cell(row=r, column=5, value=avg)
            c.number_format = fmt
            c.font = BOLD
            c.fill = LIGHT_GREEN
            # KPV range
            lo, hi = min(non_null), max(non_null)
            if isinstance(lo, (int, float)) and isinstance(hi, (int, float)):
                fmt_v = lambda v: (
                    f"{v:,.0f}" if "$" in fmt or "#,##0" == fmt else
                    f"{v:.1%}" if "%" in fmt else
                    f"{v:.1f}"
                )
                ws.cell(row=r, column=6, value=f"{fmt_v(lo)} – {fmt_v(hi)}").font = ITALIC_GREY
        r += 1

    # Per-hectare proxy section
    r += 1
    cell = ws.cell(row=r, column=1, value="PER-HECTARE PROXY (derived using approximate site areas above — sanity check only)")
    cell.font = WHITE_BOLD
    cell.fill = HEADER_BLUE
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    headers = ["Metric", "Papamoa", "Rototuna", "Waihi", "KPV avg", "KPV range"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=r, column=i, value=h)
        c.font = WHITE_BOLD
        c.fill = HEADER_BLUE
    r += 1

    ha_metrics = [
        ("Construction $/ha",       "construction_total",   "land_area_ha", "$#,##0"),
        ("Infrastructure $/ha",     None,                   "land_area_ha", "$#,##0"),  # derived from per-unit × units / ha
        ("Total Dev Cost $/ha",     "total_dev_cost",       "land_area_ha", "$#,##0"),
    ]
    for label, tot_key, ha_key, fmt in ha_metrics:
        ws.cell(row=r, column=1, value=label)
        ha = VILLAGE_RATES[ha_key]
        if tot_key:
            tot = VILLAGE_RATES[tot_key]
        else:
            # infrastructure derived
            per_u = VILLAGE_RATES["infrastructure_per_unit"]
            units = VILLAGE_RATES["units"]
            tot = tuple(p*u if p else None for p, u in zip(per_u, units))
        rates = tuple(t/h if (t and h) else None for t, h in zip(tot, ha))
        for i, v in enumerate(rates, start=2):
            c = ws.cell(row=r, column=i, value=v if v is not None else "")
            if v is not None:
                c.number_format = fmt
        non_null = [x for x in rates if x is not None]
        if non_null:
            avg = sum(non_null)/len(non_null)
            c = ws.cell(row=r, column=5, value=avg)
            c.number_format = fmt
            c.font = BOLD
            c.fill = LIGHT_GREEN
            lo, hi = min(non_null), max(non_null)
            ws.cell(row=r, column=6, value=f"${lo:,.0f} – ${hi:,.0f}").font = ITALIC_GREY
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Land areas (ha) are approximate — confirm per project. Per-ha rates vary with density and site complexity.").font = ITALIC_GREY
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)

    return r + 2

# -------- RETIREMENT G1 --------
def enrich_rv_g1():
    wb = openpyxl.load_workbook(DST_RV1)
    ws_a = wb["Assumptions"]

    # Dropdown: Land scenario at Assumptions!B8
    add_dropdown(ws_a, "B8", RV_LAND_SCENARIOS, "Land scenario",
                 "Choose one of: Ownership, Deferred, Lease, Vendor Finance")

    # Append Village Reference Rates block after the existing assumptions (find last used row)
    last_row = ws_a.max_row + 3
    new_last = add_village_rates_block(ws_a, last_row, "rv")

    # Refresh Last calibrated
    ws_a["G1"] = f"Last calibrated: {TODAY}"

    wb.save(DST_RV1)
    print(f"  RV G1 enriched: dropdown on B8, village rates block from row {last_row}")

# -------- RETIREMENT G2 --------
def enrich_rv_g2():
    wb = openpyxl.load_workbook(DST_RV2)
    ws_i = wb["Inputs"]

    # Dropdown: Land scenario at Inputs!C7
    add_dropdown(ws_i, "C7", RV_LAND_SCENARIOS, "Land scenario",
                 "Choose one of: Ownership, Deferred, Lease, Vendor Finance")

    # Add Village Reference Rates block on a new sheet (the Inputs sheet is full)
    if "Village Benchmarks" not in wb.sheetnames:
        ws_b = wb.create_sheet("Village Benchmarks", index=2)
    else:
        ws_b = wb["Village Benchmarks"]
    add_village_rates_block(ws_b, 2, "rv")
    # Widths
    for col, w in zip("ABCDEF", [40, 16, 16, 16, 16, 30]):
        ws_b.column_dimensions[col].width = w

    ws_i["G1"] = f"Last calibrated: {TODAY}"
    wb.save(DST_RV2)
    print(f"  RV G2 enriched: dropdown on Inputs!C7, new 'Village Benchmarks' sheet added")

# -------- RESIDENTIAL G1 --------
def enrich_res_g1():
    wb = openpyxl.load_workbook(DST_RES1)
    ws_a = wb["Assumptions"]

    # Dropdowns
    add_dropdown(ws_a, "B7",  RES_COUNCILS,     "Council",     "Select council")
    add_dropdown(ws_a, "B12", RES_YESNO,        "Keep dwelling", "Yes or No")
    add_dropdown(ws_a, "B13", RES_TITLE_TYPES,  "Title Type",  "Select title type")
    add_dropdown(ws_a, "B14", RES_UNIT_TYPES,   "Unit Type",   "Select unit type")
    add_dropdown(ws_a, "B15", RES_TOPOGRAPHY,   "Topography",  "Select site topography")

    # Village rates block at the bottom
    last_row = ws_a.max_row + 3
    add_village_rates_block(ws_a, last_row, "residential")
    for col, w in zip("ABCDEF", [42, 16, 16, 16, 16, 30]):
        ws_a.column_dimensions[col].width = w

    ws_a["G1"] = f"Last calibrated: {TODAY}"
    wb.save(DST_RES1)
    print(f"  Residential G1 enriched: 5 dropdowns + village rates block")

# -------- RESIDENTIAL G2 --------
def enrich_res_g2():
    wb = openpyxl.load_workbook(DST_RES2)
    ws_a = wb["Assumptions"]

    # Same dropdowns as G1 (Stage 3 has same fields)
    # Find rows by scanning
    for row in ws_a.iter_rows(min_row=1, max_row=ws_a.max_row, values_only=False):
        for cell in row:
            if isinstance(cell.value, str):
                label = cell.value.strip()
                target = ws_a.cell(row=cell.row, column=2)
                target_coord = target.coordinate
                if label == "Council":
                    add_dropdown(ws_a, target_coord, RES_COUNCILS, "Council", "Select council")
                elif label == "Keep Existing Dwelling?":
                    add_dropdown(ws_a, target_coord, RES_YESNO, "Keep dwelling", "Yes or No")
                elif label == "Title Type":
                    add_dropdown(ws_a, target_coord, RES_TITLE_TYPES, "Title Type", "Select title type")
                elif label == "Unit Type":
                    add_dropdown(ws_a, target_coord, RES_UNIT_TYPES, "Unit Type", "Select unit type")
                elif label == "Site Topography":
                    add_dropdown(ws_a, target_coord, RES_TOPOGRAPHY, "Topography", "Select topography")

    last_row = ws_a.max_row + 3
    add_village_rates_block(ws_a, last_row, "residential")
    for col, w in zip("ABCDEF", [42, 16, 16, 16, 16, 30]):
        ws_a.column_dimensions[col].width = w

    ws_a["G1"] = f"Last calibrated: {TODAY}"
    wb.save(DST_RES2)
    print(f"  Residential G2 enriched: scanned + dropdowns added + village rates block")

if __name__ == "__main__":
    enrich_rv_g1()
    enrich_rv_g2()
    enrich_res_g1()
    enrich_res_g2()
    print("\nAll 4 workbooks enriched.")
