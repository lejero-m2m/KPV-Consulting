"""
Build all 4 KPV gateway feasibility workbooks from source templates.
- Residential G1 / G2: copy Stage 2 / Stage 3 templates, rebrand LEJERO -> KPV, expand councils
- Retirement G1: copy RV_Tier1, insert Assumptions sheet, refactor Gateway formulas
- Retirement G2: copy RV_Tier2, KPV-brand cover (Inputs sheet already serves as assumptions)
"""
import shutil
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy
from pathlib import Path
import datetime as dt

ROOT = Path("/home/kyle/KPV-Consulting/KPV Feasibilities")
SRC_RV1 = ROOT / "skill-reference/RV_Tier1_Gateway_Feasibility.xlsx"
SRC_RV2 = ROOT / "skill-reference/RV_Tier2_Comprehensive_Feasibility.xlsx"
SRC_RES1 = ROOT / "examples/Stage2_Gateway_Feasibility_example Template.xlsx"
SRC_RES2 = ROOT / "examples/Stage3_Full_Feasibility_example.xlsx"

DST_RV1 = ROOT / "KPV Retirement Gateway 1.xlsx"
DST_RV2 = ROOT / "KPV Retirement Gateway 2.xlsx"
DST_RES1 = ROOT / "KPV Residential Gateway 1.xlsx"
DST_RES2 = ROOT / "KPV Residential Gateway 2.xlsx"

YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
LIGHT_BLUE = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")
HEADER_BLUE = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
TITLE_BLUE = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
WHITE_BOLD = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(color="FFFFFF", bold=True, size=14)
BLUE_INPUT = Font(color="0070C0")

# -------- RESIDENTIAL G1 (from Stage 2) --------
def build_res_g1():
    shutil.copy(SRC_RES1, DST_RES1)
    wb = openpyxl.load_workbook(DST_RES1)

    # Rebrand: LEJERO -> KPV throughout Assumptions header and Cover
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    if "LEJERO" in cell.value:
                        cell.value = cell.value.replace("LEJERO", "KPV")
                    if "Lejero Standard" in cell.value:
                        cell.value = cell.value.replace("Lejero Standard", "KPV Standard")
                    if "Lejero" in cell.value:
                        cell.value = cell.value.replace("Lejero", "KPV")

    # Expand council notes to include KPV portfolio councils
    ws_a = wb["Assumptions"]
    if ws_a["E7"].value and "TCC" in str(ws_a["E7"].value):
        ws_a["E7"].value = "TCC / WBOPDC / Matamata-Piako DC / SDC / CODC / Other"
    if ws_a["E57"].value and "TCC" in str(ws_a["E57"].value):
        ws_a["E57"].value = "TCC ~$38k | WBOPDC ~$12k | Mat-Piako ~$7.6k | SDC TBC | CODC TBC"

    # Yellow fill on default values (column B where source is Benchmark/Default/Lejero Standard)
    for row in ws_a.iter_rows(min_row=5, max_row=70):
        b_cell, c_cell, d_cell = row[1], row[2], row[3]  # B, C, D
        if c_cell.value and str(c_cell.value) in ("Benchmark", "KPV Standard", "Council Schedule", "Council", "LINZ", "NZ Tax"):
            if b_cell.value is not None:
                b_cell.fill = YELLOW
        if c_cell.value == "User Input" and b_cell.value is not None:
            b_cell.font = BLUE_INPUT

    # Add Last Calibrated stamp
    ws_a["G1"] = f"Last calibrated: {dt.date.today().isoformat()}"
    ws_a["G1"].font = Font(italic=True, color="595959")

    wb.save(DST_RES1)
    return DST_RES1

# -------- RESIDENTIAL G2 (from Stage 3) --------
def build_res_g2():
    shutil.copy(SRC_RES2, DST_RES2)
    wb = openpyxl.load_workbook(DST_RES2)

    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    if "LEJERO" in cell.value:
                        cell.value = cell.value.replace("LEJERO", "KPV")
                    if "Lejero Standard" in cell.value:
                        cell.value = cell.value.replace("Lejero Standard", "KPV Standard")
                    if "Lejero" in cell.value:
                        cell.value = cell.value.replace("Lejero", "KPV")

    ws_a = wb["Assumptions"]
    # Yellow fill defaults + blue user inputs (same logic as G1)
    for row in ws_a.iter_rows(min_row=5, max_row=83):
        if len(row) < 4:
            continue
        b_cell, c_cell, d_cell = row[1], row[2], row[3]
        if c_cell.value and str(c_cell.value) in ("Benchmark", "KPV Standard", "Council Schedule", "Council", "LINZ", "NZ Tax"):
            if b_cell.value is not None:
                b_cell.fill = YELLOW
        if c_cell.value == "User Input" and b_cell.value is not None:
            b_cell.font = BLUE_INPUT

    ws_a["G1"] = f"Last calibrated: {dt.date.today().isoformat()}"
    ws_a["G1"].font = Font(italic=True, color="595959")

    wb.save(DST_RES2)
    return DST_RES2

# -------- RETIREMENT G1 (from RV_Tier1) — needs Assumptions sheet inserted --------
def build_rv_g1():
    shutil.copy(SRC_RV1, DST_RV1)
    wb = openpyxl.load_workbook(DST_RV1)
    ws_g = wb["Gateway"]

    # Build the Assumptions sheet
    ws_a = wb.create_sheet("Assumptions", index=1)  # between Gateway and Notes

    # Title
    ws_a["A1"] = "KPV  |  Tier 1 Gateway Feasibility  |  Retirement Village Assumptions"
    ws_a["A1"].font = TITLE_FONT
    ws_a["A1"].fill = TITLE_BLUE
    ws_a.merge_cells("A1:F1")
    ws_a.row_dimensions[1].height = 22

    ws_a["A3"] = "Blue = your project value. Yellow = KPV default (editable). Every Gateway calc references this sheet."
    ws_a["A3"].font = Font(italic=True, color="595959")
    ws_a.merge_cells("A3:F3")

    # Header row
    headers = ["Assumption", "Value", "Source", "Type", "Notes"]
    for i, h in enumerate(headers, start=1):
        cell = ws_a.cell(row=4, column=i, value=h)
        cell.font = WHITE_BOLD
        cell.fill = HEADER_BLUE

    # Section + rows: (section_header, [(label, value, source, type, note, mapped_cell_in_gateway)])
    # mapped_cell is where the OLD Gateway sheet held this value; we'll write a formula in Gateway pointing here
    sections = [
        ("SITE", [
            ("Land area (m²)",                    49109,    "User Input", "Override",    "Total site area",                  "C7"),
            ("Land price (NZD, excl GST)",        7366350,  "User Input", "Override",    "Excl GST",                         "C8"),
            ("Land scenario",                     "Ownership", "User Input", "Override", "Ownership / Deferred / Lease / Vendor Finance", "C9"),
        ]),
        ("UNITS", [
            ("Unit count",                        105,      "User Input", "Override",    "Total villa unit count",           "C12"),
            ("Average ORA price per unit (incl GST)", 859722, "User Input", "Override",  "Mid-range market",                 "C13"),
            ("Average unit size (m²)",            130,      "User Input", "Override",    "Floor area over framing",          "C14"),
        ]),
        ("DEVELOPMENT COSTS (itemised, incl GST)", [
            ("Build rate per m² (incl GST)",      3718,     "Benchmark",  "Default",     "Lake Dunstan / 107 Papamoa range $3,400-$4,200", "C17"),
            ("Site works / land development total", 14281000, "User Input", "Override",  "Civils + earthworks + services",   "C18"),
            ("Community centre (total)",          8430000,  "User Input", "Override",    "Pavilion, BBQ, communal facilities", "C19"),
            ("Preliminary costs",                 4465000,  "User Input", "Override",    "Consents, design, prelims",        "C20"),
            ("Holding, marketing, insurance, rates", 2633000, "User Input", "Override",  "Over dev period",                  "C21"),
            ("KPV management & transaction fees", 4293000,  "User Input", "Override",    "KPV operator fees total",          "C22"),
        ]),
        ("TIMING", [
            ("Development duration (years)",      6,        "User Input", "Override",    "Years from start to last unit complete", "C25"),
            ("Sales velocity (units/month)",      1.5,      "User Input", "Override",    "Steady-state absorption",          "C26"),
        ]),
        ("VENDOR FINANCE (only if Land scenario = Vendor Finance)", [
            ("Vendor interest rate (annual)",     0.04,     "Benchmark",  "Default",     "Henley benchmark 4%",              "C29"),
            ("Long Stop Date (years)",            7,        "Benchmark",  "Default",     "5-7 yrs typical",                  "C30"),
        ]),
        ("KPV RV DEFAULTS (override if needed)", [
            ("DMF percentage (on resale price)",  0.24,     "Benchmark",  "Default",     "NZ industry typical 24%",          "C33"),
            ("Average resident tenure (years)",   8,        "Benchmark",  "Default",     "ILU industry avg 7-9 yrs",         "C34"),
            ("Resale fee",                        0.01,     "Benchmark",  "Default",     "NZ RV industry typical 1%",        "C35"),
            ("Capital growth (annual)",           0.02,     "Benchmark",  "Default",     "Conservative",                     "C36"),
            ("Cost inflation (annual)",           0.02,     "Benchmark",  "Default",     "Matched to capital growth",        "C37"),
            ("Discount rate (terminal value)",    0.15,     "Benchmark",  "Default",     "Stabilised RV asset",              "C38"),
            ("Preferred return on capital (A-Class)", 0.08, "Benchmark",  "Default",     "JV preferred coupon",              "C39"),
        ]),
        ("VERDICT THRESHOLDS (KPV Standard - fixed)", [
            ("PROCEED margin threshold",          0.15,     "KPV Standard", "Fixed",     "Skill spec; ≥15% AND IRR ≥15% to PROCEED", None),
            ("PROCEED IRR threshold",             0.15,     "KPV Standard", "Fixed",     "Skill spec",                       None),
            ("DECLINE margin threshold",          0.08,     "KPV Standard", "Fixed",     "Skill spec; <8% margin = DECLINE", None),
            ("DECLINE IRR threshold",             0.10,     "KPV Standard", "Fixed",     "Skill spec; <10% IRR = DECLINE",   None),
        ]),
    ]

    # Write sections and build the mapping: gateway_cell -> assumptions_cell
    gateway_to_assn = {}
    r = 5
    for section_label, items in sections:
        ws_a.cell(row=r, column=1, value=section_label).font = WHITE_BOLD
        ws_a.cell(row=r, column=1).fill = HEADER_BLUE
        ws_a.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        r += 1
        for label, value, source, ctype, note, gw_cell in items:
            ws_a.cell(row=r, column=1, value=label)
            v = ws_a.cell(row=r, column=2, value=value)
            ws_a.cell(row=r, column=3, value=source)
            ws_a.cell(row=r, column=4, value=ctype)
            ws_a.cell(row=r, column=5, value=note)
            if source == "User Input":
                v.font = BLUE_INPUT
            elif source in ("Benchmark", "KPV Standard"):
                v.fill = YELLOW
            # number format
            if isinstance(value, float) and 0 < value < 1:
                v.number_format = "0.00%"
            elif isinstance(value, (int, float)) and value > 100:
                v.number_format = "#,##0"
            if gw_cell:
                gateway_to_assn[gw_cell] = f"Assumptions!B{r}"
            r += 1

    # Column widths
    for col, w in zip("ABCDE", [42, 14, 14, 12, 60]):
        ws_a.column_dimensions[col].width = w

    # Freeze top
    ws_a.freeze_panes = "A5"

    # Last calibrated
    ws_a["G1"] = f"Last calibrated: {dt.date.today().isoformat()}"
    ws_a["G1"].font = Font(italic=True, color="595959")

    # Now update Gateway sheet: replace input cells with formulas pointing to Assumptions
    # AND update all formulas that reference those input cells
    # Strategy: for each gateway input cell, replace its value with a formula =Assumptions!Bx
    for gw_cell, assn_ref in gateway_to_assn.items():
        cell = ws_g[gw_cell]
        cell.value = f"={assn_ref}"
        cell.font = BLUE_INPUT

    # Sync verdict thresholds — Gateway G38 (verdict formula) needs to reference Assumptions thresholds
    # Original: =IF(AND(G32>=0.15,G34>=0.15),"PROCEED",IF(OR(G32<0.08,G34<0.1),"DECLINE","PROCEED WITH CAUTION"))
    # Find threshold cells we just wrote — easiest: assume the last 4 are thresholds in the order written
    # We wrote PROCEED margin, PROCEED IRR, DECLINE margin, DECLINE IRR — find their rows
    threshold_rows = []
    for row in ws_a.iter_rows(min_row=5):
        a = row[0].value
        if a in ("PROCEED margin threshold", "PROCEED IRR threshold", "DECLINE margin threshold", "DECLINE IRR threshold"):
            threshold_rows.append(row[1].coordinate)
    if len(threshold_rows) == 4:
        pm, pi, dm, di = [f"Assumptions!{c}" for c in threshold_rows]
        verdict_formula = f'=IF(AND(G32>={pm},G34>={pi}),"PROCEED",IF(OR(G32<{dm},G34<{di}),"DECLINE","PROCEED WITH CAUTION"))'
        # Find the cell that currently holds the verdict formula (F38 in template, may be merged)
        for r in range(36, 42):
            for col in ("F", "G", "H"):
                c = ws_g[f"{col}{r}"]
                if isinstance(c.value, str) and "PROCEED" in c.value and "DECLINE" in c.value:
                    try:
                        c.value = verdict_formula
                    except AttributeError:
                        # Merged — unmerge first
                        for mr in list(ws_g.merged_cells.ranges):
                            if c.coordinate in mr:
                                ws_g.unmerge_cells(str(mr))
                                break
                        ws_g[c.coordinate].value = verdict_formula
                    break

    # Update title to KPV-brand
    ws_g["B2"] = "KPV TIER 1 GATEWAY FEASIBILITY - Retirement Village"

    # Style: title bar
    ws_g["B2"].font = TITLE_FONT
    ws_g["B2"].fill = TITLE_BLUE

    wb.save(DST_RV1)
    return DST_RV1

# -------- RETIREMENT G2 (from RV_Tier2) — keep structure, KPV-brand, yellow-fill defaults --------
def build_rv_g2():
    shutil.copy(SRC_RV2, DST_RV2)
    wb = openpyxl.load_workbook(DST_RV2)
    ws_i = wb["Inputs"]

    # KPV-brand the title
    ws_i["B2"] = "KPV TIER 2 COMPREHENSIVE FEASIBILITY - Retirement Village (Inputs & Assumptions)"
    ws_i["B2"].font = TITLE_FONT
    ws_i["B2"].fill = TITLE_BLUE

    # The "KPV DEFAULTS" section starts at row 41 — yellow-fill those defaults (C42-C49)
    default_rows = [42, 43, 44, 45, 46, 47, 48, 49, 22]  # +vendor interest 22
    for r in default_rows:
        cell = ws_i.cell(row=r, column=3)
        if cell.value is not None:
            cell.fill = YELLOW

    # Blue-input convention on user-input cells (project-specific): rows 5,6,7,10,11,12,13,16-22,25,26,27,30-35,38,39
    user_rows = [5, 6, 7, 10, 11, 12, 13, 16, 17, 18, 19, 20, 21, 25, 26, 27, 30, 31, 32, 33, 34, 35, 38, 39]
    for r in user_rows:
        cell = ws_i.cell(row=r, column=3)
        if cell.value is not None and r not in default_rows:
            if not isinstance(cell.value, str) or not cell.value.startswith("="):
                cell.font = BLUE_INPUT

    # Add Last Calibrated stamp
    ws_i["G1"] = f"Last calibrated: {dt.date.today().isoformat()}"
    ws_i["G1"].font = Font(italic=True, color="595959")

    wb.save(DST_RV2)
    return DST_RV2

# -------- Run --------
if __name__ == "__main__":
    results = []
    for name, fn in [("Residential G1", build_res_g1), ("Residential G2", build_res_g2),
                     ("Retirement G1", build_rv_g1), ("Retirement G2", build_rv_g2)]:
        path = fn()
        results.append((name, path, path.stat().st_size))
        print(f"OK {name}: {path.name}  ({path.stat().st_size:,} bytes)")
    print(f"\nTotal: {len(results)} workbooks built.")
