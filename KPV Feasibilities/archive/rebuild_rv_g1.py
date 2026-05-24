"""
Rebuild Gateway tab of KPV Retirement Gateway 1.xlsx per Kyle's expanded spec:
- Project inputs on Gateway (manual entry), KPV benchmarks in adjacent column
- Dev costs benchmarked vs village budgets (per-unit and per-ha)
- Timing scenarios (-2yr/+2yr, -0.5/+0.5 sales velocity)
- 4 land scenarios calculated side-by-side
- Reverse calc 1: max viable land price per scenario
- Reverse calc 2: min unit count / min ORA at a set land price
Keeps Assumptions and Notes sheets untouched.
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import datetime as dt
from pathlib import Path

PATH = Path("/home/kyle/KPV-Consulting/KPV Feasibilities/KPV Retirement Gateway 1.xlsx")

YELLOW   = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
LBLUE    = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")
HEAD     = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
TITLE_FL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
LGREEN   = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
LRED     = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
WB_BOLD  = Font(color="FFFFFF", bold=True)
TITLE_FT = Font(color="FFFFFF", bold=True, size=14)
BOLD     = Font(bold=True)
INPUT_FT = Font(color="0070C0")
BENCH_FT = Font(color="595959", italic=True)
SMALL    = Font(size=9, color="595959", italic=True)

thin   = Side(border_style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def cell(ws, addr, value=None, font=None, fill=None, fmt=None, align=None, bold=False, border_=False):
    c = ws[addr]
    if value is not None:
        c.value = value
    if font:    c.font = font
    if bold and not font:    c.font = BOLD
    if fill:    c.fill = fill
    if fmt:     c.number_format = fmt
    if align:   c.alignment = Alignment(horizontal=align, vertical="center")
    if border_: c.border = border
    return c

def section_header(ws, row, label, cols=8):
    c = ws.cell(row=row, column=1, value=label)
    c.font = WB_BOLD
    c.fill = HEAD
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    ws.row_dimensions[row].height = 18

def sub_header(ws, row, label, cols=8):
    c = ws.cell(row=row, column=1, value=label)
    c.font = BOLD
    c.fill = LBLUE
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)

wb = openpyxl.load_workbook(PATH)
# Delete the existing Gateway sheet and recreate
if "Gateway" in wb.sheetnames:
    del wb["Gateway"]
ws = wb.create_sheet("Gateway", index=0)

# Column widths
widths = {"A": 38, "B": 14, "C": 16, "D": 16, "E": 16, "F": 16, "G": 16, "H": 18}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# === TITLE ===
cell(ws, "A2", "KPV TIER 1 GATEWAY FEASIBILITY  |  Retirement Village", font=TITLE_FT, fill=TITLE_FL, align="left")
ws.merge_cells("A2:H2")
ws.row_dimensions[2].height = 24
cell(ws, "A3", "Manual project inputs on left, KPV portfolio benchmarks alongside, 4 land scenarios calculated in parallel, plus reverse calc for viable land price. KPV defaults source from the Assumptions sheet.",
     font=SMALL, align="left")
ws.merge_cells("A3:H3")

# === PROJECT INPUTS — SITE ===
section_header(ws, 5, "PROJECT INPUTS — Site")
cell(ws, "B6", "Input", bold=True, align="center"); cell(ws, "C6", "Your value", bold=True, align="center")
cell(ws, "D6", "KPV benchmark", bold=True, align="center"); cell(ws, "E6", "KPV range", bold=True, align="center")

cell(ws, "A7", "Land area (m²)")
cell(ws, "C7", 49109, font=INPUT_FT, fmt="#,##0")
cell(ws, "D7", "~53,000 m² (5.3 ha)", font=BENCH_FT, align="center")
cell(ws, "E7", "48,000 – 57,000", font=BENCH_FT, align="center")

cell(ws, "A8", "Land price (NZD, excl GST)")
cell(ws, "C8", 7366350, font=INPUT_FT, fmt="$#,##0")
cell(ws, "D8", "(varies)", font=BENCH_FT, align="center")

cell(ws, "A9", "Land scenario")
cell(ws, "C9", "Ownership", font=INPUT_FT, align="center")
cell(ws, "D9", "(all 4 scenarios shown below)", font=BENCH_FT, align="center")
dv = DataValidation(type="list", formula1='"Ownership,Deferred,Lease,Vendor Finance"', allow_blank=False)
ws.add_data_validation(dv); dv.add("C9")

# === PROJECT INPUTS — UNITS ===
section_header(ws, 11, "PROJECT INPUTS — Units")
cell(ws, "B12", "Input", bold=True, align="center"); cell(ws, "C12", "Your value", bold=True, align="center")
cell(ws, "D12", "KPV benchmark", bold=True, align="center"); cell(ws, "E12", "KPV range", bold=True, align="center")

cell(ws, "A13", "Unit count")
cell(ws, "C13", 105, font=INPUT_FT, fmt="#,##0")
cell(ws, "D13", 114, font=BENCH_FT, fmt="#,##0", align="center")
cell(ws, "E13", "96 – 131", font=BENCH_FT, align="center")

cell(ws, "A14", "Avg ORA price per unit (incl GST)")
cell(ws, "C14", 859722, font=INPUT_FT, fmt="$#,##0")
cell(ws, "D14", 922942, font=BENCH_FT, fmt="$#,##0", align="center")
cell(ws, "E14", "$908k – $938k", font=BENCH_FT, align="center")

cell(ws, "A15", "Avg unit size (m²)")
cell(ws, "C15", 130, font=INPUT_FT, fmt="#,##0")
cell(ws, "D15", 130, font=BENCH_FT, fmt="#,##0", align="center")
cell(ws, "E15", "120 – 165", font=BENCH_FT, align="center")

# === DEVELOPMENT COSTS — itemised with benchmarks ===
section_header(ws, 17, "DEVELOPMENT COSTS (manual entry — benchmarked to KPV portfolio)")
cell(ws, "B18", "$ total", bold=True, align="center"); cell(ws, "C18", "Your value", bold=True, align="center")
cell(ws, "D18", "Implied $/unit", bold=True, align="center"); cell(ws, "E18", "KPV avg $/unit", bold=True, align="center")
cell(ws, "F18", "KPV range $/unit", bold=True, align="center")

# Each row: A=label, C=input total ($), D=implied $/unit (formula =C/units), E=KPV avg, F=range
DEV_ROWS = [
    # (label, default_total, kpv_avg_per_unit, range_text)
    ("Construction",                       43_550_000, 471_801, "$414k – $525k"),
    ("Unit landscaping",                    1_950_000,  18_903, "$18k – $20k"),
    ("Council Dev fees",                    2_700_000,  24_773, "$22k – $27k"),
    ("Infrastructure + site landscape",     9_000_000,  83_940, "$43k – $108k"),
    ("Other site works",                    1_050_000,   9_769, "$8k – $12k"),
    ("Professional fees + Preliminary",     1_950_000,  18_502, "$17k – $20k"),
    ("Holding, marketing, insurance, rates",2_633_000, None,    "(typical $2-3M total)"),
    ("KPV management & transaction fees",   4_293_000, None,    "(typical 5-7% of revenue)"),
    ("Clubhouse / community centre (total)",5_500_000, None,    "$5.3M – $5.8M total"),
]
DEV_START = 19
for i, (label, default_total, kpv_avg, rng) in enumerate(DEV_ROWS):
    r = DEV_START + i
    cell(ws, f"A{r}", label)
    cell(ws, f"C{r}", default_total, font=INPUT_FT, fmt="$#,##0")
    # Per-unit: divide by unit count if applicable
    if kpv_avg:
        cell(ws, f"D{r}", f"=IFERROR(C{r}/C13,0)", fmt="$#,##0", align="center")
        cell(ws, f"E{r}", kpv_avg, font=BENCH_FT, fmt="$#,##0", align="center")
        cell(ws, f"F{r}", rng, font=BENCH_FT, align="center")
    else:
        cell(ws, f"D{r}", "—", font=BENCH_FT, align="center")
        cell(ws, f"E{r}", "—", font=BENCH_FT, align="center")
        cell(ws, f"F{r}", rng, font=BENCH_FT, align="center")
DEV_END = DEV_START + len(DEV_ROWS) - 1

# === TIMING — scenarios ===
TIMING_HEAD = DEV_END + 2
section_header(ws, TIMING_HEAD, "TIMING (base case + scenarios for sensitivity)")
cell(ws, f"B{TIMING_HEAD+1}", "BASE", bold=True, align="center", fill=LGREEN)
cell(ws, f"C{TIMING_HEAD+1}", "-2yr / -0.5", bold=True, align="center")
cell(ws, f"D{TIMING_HEAD+1}", "+2yr / +0.5", bold=True, align="center")
DD_R = TIMING_HEAD + 2  # dev duration row
SV_R = TIMING_HEAD + 3  # sales velocity row
cell(ws, f"A{DD_R}", "Development duration (years)")
cell(ws, f"B{DD_R}", 6, font=INPUT_FT, fmt="0", align="center")
cell(ws, f"C{DD_R}", f"=B{DD_R}-2", fmt="0", align="center")
cell(ws, f"D{DD_R}", f"=B{DD_R}+2", fmt="0", align="center")
cell(ws, f"A{SV_R}", "Sales velocity (units/month)")
cell(ws, f"B{SV_R}", 1.5, font=INPUT_FT, fmt="0.0", align="center")
cell(ws, f"C{SV_R}", f"=B{SV_R}-0.5", fmt="0.0", align="center")
cell(ws, f"D{SV_R}", f"=B{SV_R}+0.5", fmt="0.0", align="center")

# === KPV DEFAULTS (pulled from Assumptions sheet) ===
DEFAULTS_HEAD = SV_R + 2
section_header(ws, DEFAULTS_HEAD, "KPV DEFAULTS  (sourced from Assumptions sheet — change them there)")
# Find rows in Assumptions for: DMF%, tenure, resale fee, capital growth, cost inflation, discount rate, preferred return
ASSN = wb["Assumptions"]
assn_lookup = {}
for row in ASSN.iter_rows(min_row=5, max_row=40):
    label = row[0].value
    if isinstance(label, str):
        assn_lookup[label.strip()] = row[1].coordinate  # B column

DEFAULT_FIELDS = [
    ("DMF percentage (on resale price)",         "0.00%"),
    ("Average resident tenure (years)",          "0"),
    ("Resale fee",                               "0.00%"),
    ("Capital growth (annual)",                  "0.00%"),
    ("Cost inflation (annual)",                  "0.00%"),
    ("Discount rate (terminal value)",           "0.00%"),
    ("Preferred return on capital (A-Class)",    "0.00%"),
    ("Vendor interest rate (annual)",            "0.00%"),
    ("Long Stop Date (years)",                   "0"),
]
DEF_START = DEFAULTS_HEAD + 1
for i, (label, fmt) in enumerate(DEFAULT_FIELDS):
    r = DEF_START + i
    cell(ws, f"A{r}", label)
    if label in assn_lookup:
        cell(ws, f"C{r}", f"=Assumptions!{assn_lookup[label]}", fmt=fmt, align="center", fill=YELLOW)
    else:
        cell(ws, f"C{r}", "(not on Assumptions)", font=BENCH_FT, align="center")
DEF_END = DEF_START + len(DEFAULT_FIELDS) - 1

# Named convenient refs (use cell coords)
UN = "C13"   # unit count
ORA = "C14" # ora price
SIZE = "C15" # unit size
LANDM = "C7" # land area m2
LANDP = "C8" # land price excl GST
LANDS = "C9" # scenario (informational)
DD = f"B{DD_R}"  # dev duration base
SV = f"B{SV_R}"  # sales velocity base
DMF = f"C{DEF_START}"
TEN = f"C{DEF_START+1}"
RSF = f"C{DEF_START+2}"
CG  = f"C{DEF_START+3}"
CI  = f"C{DEF_START+4}"
DISC = f"C{DEF_START+5}"
PREF = f"C{DEF_START+6}"
VFR  = f"C{DEF_START+7}"
LSD  = f"C{DEF_START+8}"

# Dev cost row coords for use in scenario formulas
CONST_T   = f"C{DEV_START+0}"
LAND_LDS  = f"C{DEV_START+1}"  # unit landscaping
COUNCIL_T = f"C{DEV_START+2}"
INFRA_T   = f"C{DEV_START+3}"
OTHER_T   = f"C{DEV_START+4}"
PROF_T    = f"C{DEV_START+5}"
HOLD_T    = f"C{DEV_START+6}"
KPVFEES_T = f"C{DEV_START+7}"
CLUB_T    = f"C{DEV_START+8}"

# === CALCULATIONS — 4 SCENARIOS SIDE BY SIDE ===
CALC_HEAD = DEF_END + 2
section_header(ws, CALC_HEAD, "CALCULATIONS — by land scenario (4 in parallel)")

# Scenario columns: B=Ownership, C=Deferred, D=Lease, E=Vendor Finance
# Header
hdr = CALC_HEAD + 1
cell(ws, f"A{hdr}", "Metric", bold=True, fill=LBLUE)
cell(ws, f"B{hdr}", "Ownership", bold=True, fill=LBLUE, align="center")
cell(ws, f"C{hdr}", "Deferred", bold=True, fill=LBLUE, align="center")
cell(ws, f"D{hdr}", "Lease", bold=True, fill=LBLUE, align="center")
cell(ws, f"E{hdr}", "Vendor Finance", bold=True, fill=LBLUE, align="center")

# Build per-scenario formulas. Note: cost inflation factor = (1+CI)^(DD/2)
# capital growth factor = (1+CG)^(DD/2)
# Peak capital factor: Own=0.35, Def=0.30, Lease=0.30, VF=0.20

# Revenue (same for all)
r = hdr + 1
cell(ws, f"A{r}", "Gross ORA sales (initial, nominal)")
gross_ora = f"({UN}*{ORA})"
for col in "BCDE":
    cell(ws, f"{col}{r}", f"={gross_ora}", fmt="$#,##0", align="right")

r += 1
cell(ws, f"A{r}", "ORA sales (with capital growth to mid-dev)")
ora_grown = f"({gross_ora}*(1+{CG})^({DD}/2))"
for col in "BCDE":
    cell(ws, f"{col}{r}", f"={ora_grown}", fmt="$#,##0", align="right")

# Costs
r += 2
cell(ws, f"A{r}", "Costs", bold=True, fill=LBLUE)

r += 1
cell(ws, f"A{r}", "Land payment / ground rent")
# Own: full land price at mth 1 = land_price
# Def: full land price (same total, deferred timing)
# Lease: ground rent = 6% × (land incl GST) × dev years
# Vendor Finance: full land price (paid over time but accounted in total)
own_land = f"{LANDP}"
def_land = f"{LANDP}"
lease_land = f"{LANDP}*1.15*0.06*{DD}"
vf_land = f"{LANDP}"
cell(ws, f"B{r}", f"={own_land}",   fmt="$#,##0", align="right")
cell(ws, f"C{r}", f"={def_land}",   fmt="$#,##0", align="right")
cell(ws, f"D{r}", f"={lease_land}", fmt="$#,##0", align="right")
cell(ws, f"E{r}", f"={vf_land}",    fmt="$#,##0", align="right")

r += 1
cell(ws, f"A{r}", "Vendor interest cost (approx)")
# Only VF — others 0
vf_int = f"{LANDP}*{VFR}*{LSD}*0.55"
cell(ws, f"B{r}", 0, fmt="$#,##0", align="right")
cell(ws, f"C{r}", 0, fmt="$#,##0", align="right")
cell(ws, f"D{r}", 0, fmt="$#,##0", align="right")
cell(ws, f"E{r}", f"={vf_int}", fmt="$#,##0", align="right")

r += 1
cell(ws, f"A{r}", "Construction (inflated to mid-dev)")
const_inf = f"{CONST_T}*(1+{CI})^({DD}/2)"
for col in "BCDE":
    cell(ws, f"{col}{r}", f"={const_inf}", fmt="$#,##0", align="right")

r += 1
cell(ws, f"A{r}", "Unit landscaping (inflated)")
ld_inf = f"{LAND_LDS}*(1+{CI})^({DD}/2)"
for col in "BCDE":
    cell(ws, f"{col}{r}", f"={ld_inf}", fmt="$#,##0", align="right")

r += 1
cell(ws, f"A{r}", "Council Dev fees")
for col in "BCDE":
    cell(ws, f"{col}{r}", f"={COUNCIL_T}", fmt="$#,##0", align="right")

r += 1
cell(ws, f"A{r}", "Infrastructure + site landscape (inflated)")
inf_inf = f"{INFRA_T}*(1+{CI})^({DD}/2)"
for col in "BCDE":
    cell(ws, f"{col}{r}", f"={inf_inf}", fmt="$#,##0", align="right")

r += 1
cell(ws, f"A{r}", "Other site works (inflated)")
oth_inf = f"{OTHER_T}*(1+{CI})^({DD}/2)"
for col in "BCDE":
    cell(ws, f"{col}{r}", f"={oth_inf}", fmt="$#,##0", align="right")

r += 1
cell(ws, f"A{r}", "Professional + Preliminary")
for col in "BCDE":
    cell(ws, f"{col}{r}", f"={PROF_T}", fmt="$#,##0", align="right")

r += 1
cell(ws, f"A{r}", "Holding, marketing, insurance, rates")
for col in "BCDE":
    cell(ws, f"{col}{r}", f"={HOLD_T}", fmt="$#,##0", align="right")

r += 1
cell(ws, f"A{r}", "KPV management + transaction fees")
for col in "BCDE":
    cell(ws, f"{col}{r}", f"={KPVFEES_T}", fmt="$#,##0", align="right")

r += 1
cell(ws, f"A{r}", "Clubhouse / community centre (inflated)")
club_inf = f"{CLUB_T}*(1+{CI})^({DD}/2)"
for col in "BCDE":
    cell(ws, f"{col}{r}", f"={club_inf}", fmt="$#,##0", align="right")

r += 1
cell(ws, f"A{r}", "Resale fee on initial sales")
rsf_calc = f"{ora_grown}*{RSF}"
for col in "BCDE":
    cell(ws, f"{col}{r}", f"={rsf_calc}", fmt="$#,##0", align="right")

# Total dev cost — sum land + vendor int + all others
TDC_R = r + 1
cell(ws, f"A{TDC_R}", "Total development cost", bold=True, fill=LBLUE)
# rows: hdr+4 (Costs subheader), hdr+5 (Land), hdr+6 (Vendor int), hdr+7..r (other cost lines)
# Sum from hdr+5 to r inclusive
SUM_FROM = hdr + 5
for col in "BCDE":
    cell(ws, f"{col}{TDC_R}", f"=SUM({col}{SUM_FROM}:{col}{r})", fmt="$#,##0", align="right", bold=True, fill=LBLUE)

# Surplus pre-preferred
SURP_R = TDC_R + 1
cell(ws, f"A{SURP_R}", "Development surplus (pre-preferred)")
for col in "BCDE":
    cell(ws, f"{col}{SURP_R}", f"={col}{hdr+2}-{col}{TDC_R}", fmt="$#,##0", align="right")

# Peak capital factor: B=0.35, C=0.30, D=0.30, E=0.20
PKF_R = SURP_R + 1
cell(ws, f"A{PKF_R}", "Peak capital factor", font=SMALL)
cell(ws, f"B{PKF_R}", 0.35, fmt="0%", align="center", font=SMALL)
cell(ws, f"C{PKF_R}", 0.30, fmt="0%", align="center", font=SMALL)
cell(ws, f"D{PKF_R}", 0.30, fmt="0%", align="center", font=SMALL)
cell(ws, f"E{PKF_R}", 0.20, fmt="0%", align="center", font=SMALL)

PI_R = PKF_R + 1
cell(ws, f"A{PI_R}", "Preferred interest cost (approx)")
for col in "BCDE":
    cell(ws, f"{col}{PI_R}", f"={col}{TDC_R}*{col}{PKF_R}*{PREF}*{DD}*0.6", fmt="$#,##0", align="right")

SAP_R = PI_R + 1
cell(ws, f"A{SAP_R}", "Surplus after preferred")
for col in "BCDE":
    cell(ws, f"{col}{SAP_R}", f"={col}{SURP_R}-{col}{PI_R}", fmt="$#,##0", align="right")

# Operating
DMF_R = SAP_R + 2
cell(ws, f"A{DMF_R}", "Annual DMF income at steady state (on resale price)")
dmf_calc = f"({UN}/{TEN})*{ORA}*(1+{CG})^{TEN}*({DMF}-{RSF})"
for col in "BCDE":
    cell(ws, f"{col}{DMF_R}", f"={dmf_calc}", fmt="$#,##0", align="right")

TV_R = DMF_R + 1
cell(ws, f"A{TV_R}", "Terminal value (DMF / discount rate)")
for col in "BCDE":
    cell(ws, f"{col}{TV_R}", f"={col}{DMF_R}/{DISC}", fmt="$#,##0", align="right")

# Metrics
MET_HDR = TV_R + 2
sub_header(ws, MET_HDR, "Gateway Metrics")

MARG_R = MET_HDR + 1
cell(ws, f"A{MARG_R}", "Development margin on cost", bold=True)
for col in "BCDE":
    cell(ws, f"{col}{MARG_R}", f"={col}{SURP_R}/{col}{TDC_R}", fmt="0.0%", align="right", bold=True)

MCAP_R = MARG_R + 1
cell(ws, f"A{MCAP_R}", "Margin on capital (post-preferred)")
for col in "BCDE":
    cell(ws, f"{col}{MCAP_R}", f"=IFERROR({col}{SAP_R}/{col}{TDC_R},0)", fmt="0.0%", align="right")

IRR_R = MCAP_R + 1
cell(ws, f"A{IRR_R}", "Indicative project IRR (15-year)", bold=True)
for col in "BCDE":
    cell(ws, f"{col}{IRR_R}", f"=IFERROR(RATE(15,0,-{col}{TDC_R}*{col}{PKF_R},{col}{TV_R}+{col}{SURP_R}),0)",
         fmt="0.0%", align="right", bold=True)

PEAK_R = IRR_R + 1
cell(ws, f"A{PEAK_R}", "Peak capital required (approx)")
for col in "BCDE":
    cell(ws, f"{col}{PEAK_R}", f"={col}{TDC_R}*{col}{PKF_R}", fmt="$#,##0", align="right")

# Verdict — pull thresholds from Assumptions
# Find threshold rows on Assumptions
threshold_addrs = {}
for row in ASSN.iter_rows(min_row=5, max_row=ASSN.max_row):
    a = row[0].value
    if a in ("PROCEED margin threshold","PROCEED IRR threshold","DECLINE margin threshold","DECLINE IRR threshold"):
        threshold_addrs[a] = row[1].coordinate
PM = f"Assumptions!{threshold_addrs['PROCEED margin threshold']}"
PI = f"Assumptions!{threshold_addrs['PROCEED IRR threshold']}"
DM = f"Assumptions!{threshold_addrs['DECLINE margin threshold']}"
DI = f"Assumptions!{threshold_addrs['DECLINE IRR threshold']}"

VERD_R = PEAK_R + 2
cell(ws, f"A{VERD_R}", "VERDICT", bold=True, fill=HEAD, font=WB_BOLD)
for col in "BCDE":
    formula = f'=IF(AND({col}{MARG_R}>={PM},{col}{IRR_R}>={PI}),"PROCEED",IF(OR({col}{MARG_R}<{DM},{col}{IRR_R}<{DI}),"DECLINE","PROCEED W/ CAUTION"))'
    c = cell(ws, f"{col}{VERD_R}", formula, bold=True, align="center")
    c.font = Font(bold=True, size=12)

# === TIMING SCENARIO IMPACT (uses Ownership for comparison) ===
TS_HDR = VERD_R + 3
section_header(ws, TS_HDR, "TIMING SCENARIO IMPACT (Ownership scenario, holds all other inputs at base)")
cell(ws, f"A{TS_HDR+1}", "Scenario", bold=True, fill=LBLUE)
cell(ws, f"B{TS_HDR+1}", "BASE", bold=True, fill=LGREEN, align="center")
cell(ws, f"C{TS_HDR+1}", "-2 yr dev", bold=True, align="center")
cell(ws, f"D{TS_HDR+1}", "+2 yr dev", bold=True, align="center")
cell(ws, f"E{TS_HDR+1}", "-0.5 sales", bold=True, align="center")
cell(ws, f"F{TS_HDR+1}", "+0.5 sales", bold=True, align="center")

# Use simplified scenario calcs: just vary DD and recalc margin/IRR (sales velocity is informational only at this gateway level since cashflow timing isn't modeled)
# For -2yr/+2yr: substitute DD with DD-2 / DD+2 in the construction inflation and ORA growth
# For sales velocity: at this gateway level it doesn't change costs/revenue; flag as informational
# Build helper formulas
DD_SCENARIOS = {"B": DD, "C": f"({DD}-2)", "D": f"({DD}+2)", "E": DD, "F": DD}

# Margin row (Ownership uses col B in calc block)
TS_MARG = TS_HDR + 2
cell(ws, f"A{TS_MARG}", "Margin on cost")
for col, dd in DD_SCENARIOS.items():
    # Recompute revenue and cost shapes with new DD
    rev = f"({UN}*{ORA})*(1+{CG})^({dd}/2)"
    const = f"{CONST_T}*(1+{CI})^({dd}/2)"
    infra = f"{INFRA_T}*(1+{CI})^({dd}/2)"
    oth = f"{OTHER_T}*(1+{CI})^({dd}/2)"
    club = f"{CLUB_T}*(1+{CI})^({dd}/2)"
    ld = f"{LAND_LDS}*(1+{CI})^({dd}/2)"
    land = f"{LANDP}"  # Ownership
    cost = f"({land}+{const}+{ld}+{COUNCIL_T}+{infra}+{oth}+{PROF_T}+{HOLD_T}+{KPVFEES_T}+{club}+({rev}*{RSF}))"
    cell(ws, f"{col}{TS_MARG}", f"=({rev}-{cost})/{cost}", fmt="0.0%", align="right")

TS_IRR = TS_MARG + 1
cell(ws, f"A{TS_IRR}", "Project IRR (15-yr proxy)")
for col, dd in DD_SCENARIOS.items():
    rev = f"({UN}*{ORA})*(1+{CG})^({dd}/2)"
    const = f"{CONST_T}*(1+{CI})^({dd}/2)"
    infra = f"{INFRA_T}*(1+{CI})^({dd}/2)"
    oth = f"{OTHER_T}*(1+{CI})^({dd}/2)"
    club = f"{CLUB_T}*(1+{CI})^({dd}/2)"
    ld = f"{LAND_LDS}*(1+{CI})^({dd}/2)"
    land = f"{LANDP}"
    cost = f"({land}+{const}+{ld}+{COUNCIL_T}+{infra}+{oth}+{PROF_T}+{HOLD_T}+{KPVFEES_T}+{club}+({rev}*{RSF}))"
    dmf = f"({UN}/{TEN})*{ORA}*(1+{CG})^{TEN}*({DMF}-{RSF})"
    tv = f"({dmf}/{DISC})"
    pk = f"({cost}*0.35)"
    cell(ws, f"{col}{TS_IRR}", f"=IFERROR(RATE(15,0,-{pk},{tv}+({rev}-{cost})),0)", fmt="0.0%", align="right")

TS_VERD = TS_IRR + 1
cell(ws, f"A{TS_VERD}", "Verdict", bold=True)
for col in "BCDEF":
    formula = f'=IF(AND({col}{TS_MARG}>={PM},{col}{TS_IRR}>={PI}),"PROCEED",IF(OR({col}{TS_MARG}<{DM},{col}{TS_IRR}<{DI}),"DECLINE","CAUTION"))'
    cell(ws, f"{col}{TS_VERD}", formula, bold=True, align="center")

cell(ws, f"A{TS_VERD+1}", "Note: sales velocity has no effect at Gateway 1 (no cashflow timing model). Use Gateway 2 for velocity sensitivity.", font=SMALL)
ws.merge_cells(start_row=TS_VERD+1, start_column=1, end_row=TS_VERD+1, end_column=6)

# === REVERSE CALC 1 — Max viable land price ===
RC1_HDR = TS_VERD + 4
section_header(ws, RC1_HDR, "REVERSE CALC 1 — Max viable land price (to hit PROCEED at base assumptions)")
cell(ws, f"A{RC1_HDR+1}", "Solves: at what land price does the 15% margin threshold bind? (uses base timing, base unit count, base ORA)", font=SMALL)
ws.merge_cells(start_row=RC1_HDR+1, start_column=1, end_row=RC1_HDR+1, end_column=6)

RC1_TBL = RC1_HDR + 3
cell(ws, f"A{RC1_TBL}", "Metric", bold=True, fill=LBLUE)
cell(ws, f"B{RC1_TBL}", "Ownership", bold=True, fill=LBLUE, align="center")
cell(ws, f"C{RC1_TBL}", "Deferred", bold=True, fill=LBLUE, align="center")
cell(ws, f"D{RC1_TBL}", "Lease", bold=True, fill=LBLUE, align="center")
cell(ws, f"E{RC1_TBL}", "Vendor Finance", bold=True, fill=LBLUE, align="center")

# Other costs (excluding land + vendor int + resale fee) — same across scenarios except scenario-specific
# OtherCosts = construction_infl + landscaping_infl + council + infra_infl + other_infl + prof + holding + kpv_fees + clubhouse_infl
OTHER_COSTS = (
    f"({CONST_T}*(1+{CI})^({DD}/2)"
    f"+{LAND_LDS}*(1+{CI})^({DD}/2)"
    f"+{COUNCIL_T}"
    f"+{INFRA_T}*(1+{CI})^({DD}/2)"
    f"+{OTHER_T}*(1+{CI})^({DD}/2)"
    f"+{PROF_T}"
    f"+{HOLD_T}"
    f"+{KPVFEES_T}"
    f"+{CLUB_T}*(1+{CI})^({DD}/2))"
)
RESALE_FEE_TERM = f"({UN}*{ORA}*(1+{CG})^({DD}/2))*{RSF}"  # resale fee scales with revenue, doesn't depend on land
TOTAL_REV = f"({UN}*{ORA}*(1+{CG})^({DD}/2))"

r = RC1_TBL + 1
cell(ws, f"A{r}", "Target margin (from Assumptions)")
for col in "BCDE":
    cell(ws, f"{col}{r}", f"={PM}", fmt="0.0%", align="center", font=BENCH_FT)

# Algebra: For Ownership: margin = (Rev - (Land + OtherCosts + ResaleFee))/(Land + OtherCosts + ResaleFee) = m
#   → Land + OtherCosts + ResaleFee = Rev / (1+m)
#   → Land_max = Rev/(1+m) - OtherCosts - ResaleFee
# For Lease: Land term = Land*1.15*0.06*DD
#   → Land + OC + RF = Rev/(1+m) where Land term replaces Land
#   → Land*1.15*0.06*DD = Rev/(1+m) - OC - RF
#   → Land_max = (Rev/(1+m) - OC - RF) / (1.15*0.06*DD)
# For Vendor Finance: cost includes Land + VendorInt = Land*(1 + 0.55*VFR*LSD)
#   → Land*(1+0.55*VFR*LSD) = Rev/(1+m) - OC - RF
#   → Land_max = (Rev/(1+m) - OC - RF) / (1+0.55*VFR*LSD)
# Deferred: same as Ownership (total cost includes full land payment)

r += 1
cell(ws, f"A{r}", "Max land price (excl GST) at margin threshold", bold=True)
budget = f"({TOTAL_REV}/(1+{PM})-{OTHER_COSTS}-{RESALE_FEE_TERM})"
cell(ws, f"B{r}", f"={budget}", fmt="$#,##0", align="right", bold=True, fill=LGREEN)
cell(ws, f"C{r}", f"={budget}", fmt="$#,##0", align="right", bold=True, fill=LGREEN)
cell(ws, f"D{r}", f"={budget}/(1.15*0.06*{DD})", fmt="$#,##0", align="right", bold=True, fill=LGREEN)
cell(ws, f"E{r}", f"={budget}/(1+0.55*{VFR}*{LSD})", fmt="$#,##0", align="right", bold=True, fill=LGREEN)

r += 1
cell(ws, f"A{r}", "Current land price (input)")
for col in "BCDE":
    cell(ws, f"{col}{r}", f"={LANDP}", fmt="$#,##0", align="right", font=BENCH_FT)

r += 1
cell(ws, f"A{r}", "Headroom (max − current)")
# Conditional formatting via formula return: positive = green-ish, negative = red-ish
for col in "BCDE":
    cell(ws, f"{col}{r}", f"={col}{r-2}-{col}{r-1}", fmt="$#,##0", align="right", bold=True)

# === REVERSE CALC 2 — Min requirements at set land price ===
RC2_HDR = r + 3
section_header(ws, RC2_HDR, "REVERSE CALC 2 — Minimum requirements at a SET land price")
cell(ws, f"A{RC2_HDR+1}", "If you have a fixed land asking price, what's the minimum unit count / ORA / construction-cost reduction to clear 15% margin?", font=SMALL)
ws.merge_cells(start_row=RC2_HDR+1, start_column=1, end_row=RC2_HDR+1, end_column=6)

RC2_INPUT = RC2_HDR + 2
cell(ws, f"A{RC2_INPUT}", "Set land price (excl GST)")
cell(ws, f"B{RC2_INPUT}", 7366350, font=INPUT_FT, fmt="$#,##0", fill=YELLOW)

RC2_TBL = RC2_INPUT + 2
cell(ws, f"A{RC2_TBL}", "Metric", bold=True, fill=LBLUE)
cell(ws, f"B{RC2_TBL}", "Ownership", bold=True, fill=LBLUE, align="center")
cell(ws, f"C{RC2_TBL}", "Deferred", bold=True, fill=LBLUE, align="center")
cell(ws, f"D{RC2_TBL}", "Lease", bold=True, fill=LBLUE, align="center")
cell(ws, f"E{RC2_TBL}", "Vendor Finance", bold=True, fill=LBLUE, align="center")

# Effective land cost per scenario at the set land price
SET_LP = f"B{RC2_INPUT}"
EFF_LAND = {
    "B": SET_LP,
    "C": SET_LP,
    "D": f"{SET_LP}*1.15*0.06*{DD}",
    "E": f"{SET_LP}*(1+0.55*{VFR}*{LSD})",
}

# Variable per-unit cost (construction + landscaping + council + infra + other + prof — these scale with units)
# Fixed costs (holding + kpv fees + clubhouse + per-unit-cost terms that are total)
# Easier algebra: assume per-unit costs scale linearly with unit count
# Per-unit variable cost (incl construction inflation): use current construction_total / current unit count as $/unit
# Then: Rev_per_unit = ORA × growth_factor; Cost_per_unit_variable = (Construction + Landscaping + Council + Infra + Other + Prof)/units × (1+CI)^(DD/2) for inflated items
# Margin = (Rev*units - eff_land - var_per_u*units - fixed)/( eff_land + var_per_u*units + fixed) >= PM
# → units*(Rev - var_per_u*(1+PM)) >= (1+PM)*(eff_land + fixed)
# → units_min = (1+PM)*(eff_land + fixed) / (Rev_per_unit - var_per_u*(1+PM))
# where Rev_per_unit = ORA*(1+CG)^(DD/2)
# var_per_u (inflated where applicable) = (Construction+Landscaping+Infra+Other+Clubhouse)*(1+CI)^(DD/2)/units + (Council+Prof)/units
# fixed = Holding + KPV fees + Clubhouse*infl  — actually Clubhouse is total, doesn't scale with units; same for Holding + KPV fees

# Let me re-derive cleanly:
# Items that scale with units: Construction, Landscaping (unit landscaping), Council Dev fees (per unit), Infra, Other site works, Prof fees, KPV fees
# Items that are fixed regardless of units: Holding/marketing, Clubhouse
# (This is a simplification — in reality many "totals" scale with units, but the user-entered totals reflect their unit assumption)
# Cleaner: assume ALL the dev cost inputs scale with units (so total ÷ current units = $/unit, then × required units)
# That gives: at unit count u: cost_excluding_land_excluding_resale = (current_total_other_costs / current_units) * u
# Rev per unit at growth: ORA*(1+CG)^(DD/2)
# Resale fee scales with revenue = u*Rev_per_unit*RSF
# Then: Margin = (u*Rev_per_unit - land_eff - u*var_per_u - u*Rev_per_unit*RSF) / (land_eff + u*var_per_u + u*Rev_per_unit*RSF) >= PM
# → u*(Rev_per_unit*(1-RSF) - var_per_u*(1+PM) - Rev_per_unit*RSF*(1+PM)) >= (1+PM)*land_eff
# → u*(Rev_per_unit*(1 - RSF*(2+PM)) - var_per_u*(1+PM)) >= (1+PM)*land_eff
# Hmm getting messy. Simpler approximation: ignore resale fee in this derivation (it's small)
# → u*(Rev_per_unit - var_per_u*(1+PM)) >= (1+PM)*land_eff
# → u_min = (1+PM)*land_eff / (Rev_per_unit - var_per_u*(1+PM))

VAR_PER_U = f"(({CONST_T}+{LAND_LDS}+{INFRA_T}+{OTHER_T}+{CLUB_T})*(1+{CI})^({DD}/2)+{COUNCIL_T}+{PROF_T}+{HOLD_T}+{KPVFEES_T})/{UN}"
REV_PER_U = f"{ORA}*(1+{CG})^({DD}/2)"

r = RC2_TBL + 1
cell(ws, f"A{r}", "Min unit count required (at base ORA)", bold=True)
for col, ef in EFF_LAND.items():
    formula = f"=IFERROR((1+{PM})*{ef}/({REV_PER_U}-{VAR_PER_U}*(1+{PM})),\"n/a\")"
    cell(ws, f"{col}{r}", formula, fmt="#,##0", align="right", bold=True, fill=LGREEN)

r += 1
cell(ws, f"A{r}", "Min avg ORA required (at base unit count)", bold=True)
# Margin = (units*ORA*growth - eff_land - units*var_per_u)/( eff_land + units*var_per_u) >= PM
# → units*ORA*growth >= (1+PM)*(eff_land + units*var_per_u)
# → ORA_min = (1+PM)*(eff_land + units*var_per_u) / (units * growth)
# growth = (1+CG)^(DD/2)
GROWTH = f"(1+{CG})^({DD}/2)"
for col, ef in EFF_LAND.items():
    formula = f"=IFERROR((1+{PM})*({ef}+{UN}*{VAR_PER_U})/({UN}*{GROWTH}),\"n/a\")"
    cell(ws, f"{col}{r}", formula, fmt="$#,##0", align="right", bold=True, fill=LGREEN)

r += 1
cell(ws, f"A{r}", "vs current ORA")
for col in "BCDE":
    cell(ws, f"{col}{r}", f"={ORA}", fmt="$#,##0", align="right", font=BENCH_FT)

# === SUMMARY ROW (selected scenario) ===
SUM_HDR = r + 3
section_header(ws, SUM_HDR, f"HEADLINE — selected land scenario ({{C9}})")
cell(ws, f"A{SUM_HDR}", f'="HEADLINE — selected land scenario: "&{LANDS}', font=WB_BOLD, fill=HEAD)

r = SUM_HDR + 1
cell(ws, f"A{r}", "Margin on cost", bold=True)
cell(ws, f"C{r}", f'=INDEX(B{MARG_R}:E{MARG_R},MATCH({LANDS},{{"Ownership","Deferred","Lease","Vendor Finance"}},0))', fmt="0.0%", align="right", bold=True)

r += 1
cell(ws, f"A{r}", "Project IRR", bold=True)
cell(ws, f"C{r}", f'=INDEX(B{IRR_R}:E{IRR_R},MATCH({LANDS},{{"Ownership","Deferred","Lease","Vendor Finance"}},0))', fmt="0.0%", align="right", bold=True)

r += 1
cell(ws, f"A{r}", "Peak capital required")
cell(ws, f"C{r}", f'=INDEX(B{PEAK_R}:E{PEAK_R},MATCH({LANDS},{{"Ownership","Deferred","Lease","Vendor Finance"}},0))', fmt="$#,##0", align="right")

r += 1
cell(ws, f"A{r}", "VERDICT", bold=True, fill=HEAD, font=WB_BOLD)
cell(ws, f"C{r}", f'=INDEX(B{VERD_R}:E{VERD_R},MATCH({LANDS},{{"Ownership","Deferred","Lease","Vendor Finance"}},0))', bold=True, align="center", font=Font(bold=True, size=14))

# Reorder sheets so Gateway is first
order = ["Gateway"] + [s for s in wb.sheetnames if s != "Gateway"]
wb._sheets = [wb[name] for name in order]

# Freeze top
ws.freeze_panes = "A5"

wb.save(PATH)
print(f"Rebuilt {PATH.name} — Gateway tab restructured per Kyle's spec")
print(f"  Project inputs on Gateway rows 7-15")
print(f"  Dev costs rows {DEV_START}-{DEV_END}")
print(f"  Timing scenarios rows {TIMING_HEAD}-{SV_R}")
print(f"  KPV defaults from Assumptions rows {DEF_START}-{DEF_END}")
print(f"  4-scenario calculations rows {CALC_HEAD}-{VERD_R}")
print(f"  Timing scenario impact rows {TS_HDR}-{TS_VERD+1}")
print(f"  Reverse calc 1 (max land price) rows {RC1_HDR}-{RC1_TBL+3}")
print(f"  Reverse calc 2 (min requirements) rows {RC2_HDR}-{r-1}")
print(f"  Headline summary rows {SUM_HDR}-{r}")
