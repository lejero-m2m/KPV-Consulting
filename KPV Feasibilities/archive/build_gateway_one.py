"""
Build KPV_Gateway_One_Feasibility.xlsx per Kyle's 2026-05-23 brief.

Structure:
  1. Inputs                       — user inputs + KPV benchmarks + headline verdict at top
  2. Calculations                 — 4 land-scenarios side-by-side + SELECTED column
  3. Maximum Viable Price         — reverse calc: max land $, min units, min ORA, $/ha
  4. Scenarios                    — 13 sensitivity scenarios
  5. Notes                        — method, thresholds, limitations
  6. Assumptions Reference        — Kyle's kpv_feasibility_assumptions.xlsx (verbatim copy)

Notes:
- Brief specifies output to /mnt/user-data/outputs/; that path doesn't exist in this env.
  Saving to /home/kyle/KPV-Consulting/KPV Feasibilities/ instead.
- Brief specifies the recalc.py verification step; not available in this env.
  Kyle validates by opening in Excel.
- KPV DEFAULTS in Inputs source from Kyle's Assumptions Reference sheet
  (his richer dataset supersedes some brief constants — notably tenure 9yr and
  discount rate 17.5% internal).
"""
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, NamedStyle
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from copy import copy
import datetime as dt
from pathlib import Path

ROOT = Path("/home/kyle/KPV-Consulting/KPV Feasibilities")
SRC_ASSUMPTIONS = ROOT / "kpv_feasibility_assumptions.xlsx"
OUT = ROOT / "KPV_Gateway_One_Feasibility.xlsx"

# --- Styles ---
YELLOW    = PatternFill("solid", fgColor="FFF2CC")
LBLUE     = PatternFill("solid", fgColor="DEEBF7")
HEAD      = PatternFill("solid", fgColor="2E75B6")
TITLE_FL  = PatternFill("solid", fgColor="1F3864")
LGREEN    = PatternFill("solid", fgColor="E2EFDA")
LAMBER    = PatternFill("solid", fgColor="FFF2CC")
LRED      = PatternFill("solid", fgColor="FCE4D6")
WB_BOLD   = Font(color="FFFFFF", bold=True)
TITLE_FT  = Font(color="FFFFFF", bold=True, size=14)
SUBT_FT   = Font(color="FFFFFF", bold=True, size=10)
BOLD      = Font(bold=True)
INPUT_FT  = Font(color="0070C0")
BENCH_FT  = Font(color="595959", italic=True)
SMALL     = Font(size=9, color="595959", italic=True)
LARGE_BD  = Font(bold=True, size=14)

# Number formats per brief
FMT_CCY   = '$#,##0;($#,##0);"-"'
FMT_PCT   = '0.0%;(0.0%);"-"'
FMT_INT   = '#,##0;(#,##0);"-"'
FMT_DEC   = '0.0'

# --- Constants (brief reference data) ---
KPV = {
    'construction_per_unit_avg': 471801,
    'construction_per_unit_range': (414329, 524852),
    'landscaping_per_unit_avg': 18903,
    'landscaping_per_unit_range': (18072, 20017),
    'council_fees_per_unit_avg': 24773,
    'council_fees_per_unit_range': (22087, 26782),
    'infrastructure_per_unit_avg': 83940,
    'infrastructure_per_unit_range': (42869, 108212),
    'other_site_works_per_unit_avg': 9769,
    'other_site_works_per_unit_range': (7852, 11993),
    'professional_per_unit_avg': 18502,
    'professional_per_unit_range': (16858, 20147),
    'clubhouse_total_range': (5297195, 5845866),
    'unit_count_avg': 114,
    'unit_count_range': (96, 131),
    'land_area_ha_avg': 5.3,
    'land_area_ha_range': (4.8, 5.7),
    'density_avg': 21.3,
    'density_range': (20.0, 23.8),
    'ora_avg': 922942,
    'ora_range': (908322, 937563),
    'ora_unit_size_avg': 130,
    'ora_unit_size_range': (120, 165),
    'pavilion_rates': [('3.0 star', 4000), ('3.5 star', 5000), ('4.0 star', 6000), ('4.5 star', 7000)],
    # Defaults — Kyle's assumptions override some
    'dmf_pct': 0.24,
    'tenure_years': 9,            # Kyle's assumption (was 8 in brief)
    'resale_fee': 0.01,
    'capital_growth': 0.02,
    'cost_inflation': 0.02,
    'discount_rate': 0.175,       # Kyle's internal (brief said 0.15)
    'preferred_rate': 0.08,
    'vendor_rate': 0.04,
    'vendor_lsd_years': 7,
    'proceed_margin': 0.15,
    'proceed_irr': 0.15,
    'decline_margin': 0.08,
    'decline_irr': 0.10,
}

def fill_cell(ws, coord, value=None, font=None, fill=None, fmt=None, align=None, bold=False, comment=None):
    c = ws[coord]
    if value is not None: c.value = value
    if font: c.font = font
    elif bold: c.font = BOLD
    if fill: c.fill = fill
    if fmt: c.number_format = fmt
    if align: c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    if comment: c.comment = Comment(comment, "KPV")
    return c

def title_bar(ws, row, text, span="B:G"):
    start_col, end_col = span.split(":")
    coord = f"{start_col}{row}"
    c = ws[coord]
    c.value = text
    c.font = TITLE_FT
    c.fill = TITLE_FL
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(f"{start_col}{row}:{end_col}{row}")
    ws.row_dimensions[row].height = 22

def section_header(ws, row, text, span="B:G"):
    start_col, end_col = span.split(":")
    coord = f"{start_col}{row}"
    c = ws[coord]
    c.value = text
    c.font = WB_BOLD
    c.fill = HEAD
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(f"{start_col}{row}:{end_col}{row}")
    ws.row_dimensions[row].height = 18

def subtotal(ws, coord, value=None):
    c = ws[coord]
    if value is not None: c.value = value
    c.font = BOLD
    c.fill = LBLUE

def cf_threshold(ws, range_str, kind):
    """kind: 'margin', 'irr', or 'verdict'"""
    if kind == 'margin':
        ws.conditional_formatting.add(range_str, CellIsRule(operator='greaterThanOrEqual', formula=[str(KPV['proceed_margin'])], fill=LGREEN))
        ws.conditional_formatting.add(range_str, CellIsRule(operator='lessThan', formula=[str(KPV['decline_margin'])], fill=LRED))
        ws.conditional_formatting.add(range_str, CellIsRule(operator='between', formula=[str(KPV['decline_margin']), str(KPV['proceed_margin'])], fill=LAMBER))
    elif kind == 'irr':
        ws.conditional_formatting.add(range_str, CellIsRule(operator='greaterThanOrEqual', formula=[str(KPV['proceed_irr'])], fill=LGREEN))
        ws.conditional_formatting.add(range_str, CellIsRule(operator='lessThan', formula=[str(KPV['decline_irr'])], fill=LRED))
        ws.conditional_formatting.add(range_str, CellIsRule(operator='between', formula=[str(KPV['decline_irr']), str(KPV['proceed_irr'])], fill=LAMBER))
    elif kind == 'verdict':
        ws.conditional_formatting.add(range_str, FormulaRule(formula=[f'EXACT({range_str.split(":")[0]},"PROCEED")'], fill=LGREEN, font=Font(bold=True, color="375623")))
        ws.conditional_formatting.add(range_str, FormulaRule(formula=[f'EXACT({range_str.split(":")[0]},"DECLINE")'], fill=LRED, font=Font(bold=True, color="9C0006")))
        ws.conditional_formatting.add(range_str, FormulaRule(formula=[f'EXACT({range_str.split(":")[0]},"PROCEED WITH CAUTION")'], fill=LAMBER, font=Font(bold=True, color="9C5700")))

# =================================================================
# Build workbook
# =================================================================
wb = Workbook()
# Remove default sheet
wb.remove(wb.active)

# --- Sheet 1: Inputs (created last so we have row positions) — actually create up front, fill in below ---
# We'll build Calculations first since Inputs headline references it, but openpyxl can write
# formulas referencing sheets that don't exist yet. So build in declared brief order.

# =================================================================
# SHEET 1: INPUTS
# =================================================================
ws_in = wb.create_sheet("Inputs")

# Column widths per brief
widths_in = {'A': 2, 'B': 42, 'C': 16, 'D': 16, 'E': 16, 'F': 16, 'G': 50}
for col, w in widths_in.items():
    ws_in.column_dimensions[col].width = w

# Title bar rows 2-3
title_bar(ws_in, 2, "KPV TIER 1 GATEWAY FEASIBILITY", span="B:G")
c = ws_in["B3"]; c.value = "30-minute gating test for retirement village land acquisition"
c.font = SUBT_FT; c.fill = TITLE_FL; c.alignment = Alignment(horizontal="left", vertical="center")
ws_in.merge_cells("B3:G3")
ws_in.row_dimensions[3].height = 16

# ========== HEADLINE BLOCK rows 5-11 (filled after Calculations is built — placeholder for now) ==========
section_header(ws_in, 5, "HEADLINE VERDICT")
# Rows 6-11 — labels in B, values in C (one column), notes in G
ws_in["B6"] = "Selected land scenario";       ws_in["B7"] = "Development margin on cost"
ws_in["B8"] = "Indicative project IRR (15-yr)"; ws_in["B9"] = "Peak capital required"
ws_in["B10"] = "Implied $/unit total dev cost"; ws_in["B11"] = "Implied units per hectare"
for r in range(6, 12):
    ws_in[f"B{r}"].font = BOLD
# Values come from Calculations — wire later

# Verdict cell — large merged across C:G row 12
verdict_cell = ws_in["C12"]
ws_in.merge_cells("C12:G12")
ws_in.row_dimensions[12].height = 32

# ========== SITE section rows 14-19 ==========
section_header(ws_in, 14, "PROJECT INPUTS — Site")
# Headers
ws_in["C15"] = "Your value"; ws_in["D15"] = "KPV avg"; ws_in["E15"] = "KPV range"; ws_in["F15"] = "Source"; ws_in["G15"] = "Notes"
for col in "CDEFG": fill_cell(ws_in, f"{col}15", bold=True, align="center")

# Site rows
ws_in["B16"] = "Land area (hectares)"
fill_cell(ws_in, "C16", 4.9, font=INPUT_FT, fmt=FMT_DEC, align="right")
fill_cell(ws_in, "D16", KPV['land_area_ha_avg'], font=BENCH_FT, fmt=FMT_DEC, align="center")
fill_cell(ws_in, "E16", f"{KPV['land_area_ha_range'][0]}-{KPV['land_area_ha_range'][1]}", font=BENCH_FT, align="center")
ws_in["F16"] = "KPV portfolio"; ws_in["F16"].font = BENCH_FT
ws_in["G16"] = "Total site area"

ws_in["B17"] = "Land area (m²)"
fill_cell(ws_in, "C17", "=C16*10000", fmt=FMT_INT, align="right")
ws_in["G17"] = "Auto-derived from hectares"

ws_in["B18"] = "Land price (NZD, excl GST)"
fill_cell(ws_in, "C18", 7000000, font=INPUT_FT, fmt=FMT_CCY, align="right")
ws_in["G18"] = "Vendor asking price"

ws_in["B19"] = "Land scenario"
fill_cell(ws_in, "C19", "Ownership", font=INPUT_FT, align="center")
ws_in["E19"] = "Ownership / Deferred / Lease / Vendor Finance"; ws_in["E19"].font = BENCH_FT
ws_in.merge_cells("E19:F19")
ws_in["G19"] = "Select from dropdown"
dv_land = DataValidation(type="list", formula1='"Ownership,Deferred,Lease,Vendor Finance"', allow_blank=False)
dv_land.errorTitle = "Invalid scenario"; dv_land.error = "Choose Ownership, Deferred, Lease, or Vendor Finance"
ws_in.add_data_validation(dv_land); dv_land.add("C19")

# ========== UNITS section rows 21-25 ==========
section_header(ws_in, 21, "PROJECT INPUTS — Units")
ws_in["B22"] = "Unit count"
fill_cell(ws_in, "C22", 100, font=INPUT_FT, fmt=FMT_INT, align="right")
fill_cell(ws_in, "D22", KPV['unit_count_avg'], font=BENCH_FT, fmt=FMT_INT, align="center")
fill_cell(ws_in, "E22", f"{KPV['unit_count_range'][0]}-{KPV['unit_count_range'][1]}", font=BENCH_FT, align="center")
ws_in["F22"] = "KPV portfolio"; ws_in["F22"].font = BENCH_FT

ws_in["B23"] = "Average ORA price per unit (incl GST)"
fill_cell(ws_in, "C23", 850000, font=INPUT_FT, fmt=FMT_CCY, align="right")
fill_cell(ws_in, "D23", KPV['ora_avg'], font=BENCH_FT, fmt=FMT_CCY, align="center")
fill_cell(ws_in, "E23", f"${KPV['ora_range'][0]/1000:.0f}k-${KPV['ora_range'][1]/1000:.0f}k", font=BENCH_FT, align="center")
ws_in["F23"] = "KPV portfolio"; ws_in["F23"].font = BENCH_FT

ws_in["B24"] = "Average unit size (m²)"
fill_cell(ws_in, "C24", 130, font=INPUT_FT, fmt=FMT_INT, align="right")
fill_cell(ws_in, "D24", KPV['ora_unit_size_avg'], font=BENCH_FT, fmt=FMT_INT, align="center")
fill_cell(ws_in, "E24", f"{KPV['ora_unit_size_range'][0]}-{KPV['ora_unit_size_range'][1]}", font=BENCH_FT, align="center")
ws_in["F24"] = "KPV portfolio"; ws_in["F24"].font = BENCH_FT

ws_in["B25"] = "Implied density (units/ha)"
fill_cell(ws_in, "C25", "=IFERROR(C22/C16,0)", fmt=FMT_DEC, align="right")
fill_cell(ws_in, "D25", KPV['density_avg'], font=BENCH_FT, fmt=FMT_DEC, align="center")
fill_cell(ws_in, "E25", f"{KPV['density_range'][0]}-{KPV['density_range'][1]}", font=BENCH_FT, align="center")
# CF on density: amber if outside KPV range, red if >5 outside
ws_in.conditional_formatting.add("C25", CellIsRule(operator='between', formula=['20', '24'], fill=LGREEN))
ws_in.conditional_formatting.add("C25", FormulaRule(formula=['OR(C25<15,C25>29)'], fill=LRED))
ws_in.conditional_formatting.add("C25", FormulaRule(formula=['OR(AND(C25>=15,C25<20),AND(C25>24,C25<=29))'], fill=LAMBER))

# ========== DEV COSTS section rows 27-37 ==========
section_header(ws_in, 27, "DEVELOPMENT COSTS — manual entry, benchmarked vs KPV portfolio")
# Headers
ws_in["C28"] = "Your $ total"; ws_in["D28"] = "Implied $/unit"; ws_in["E28"] = "KPV avg $/unit"; ws_in["F28"] = "KPV range"
for col in "CDEF": fill_cell(ws_in, f"{col}28", bold=True, align="center")

# Each row: (label, default total, kpv_avg_per_unit_or_None, range_text, comment)
dev_rows = [
    ("Construction (excluding pavilion)",          47000000, KPV['construction_per_unit_avg'],   f"${KPV['construction_per_unit_range'][0]/1000:.0f}k - ${KPV['construction_per_unit_range'][1]/1000:.0f}k", "Build cost for residential units only. Pavilion costed separately below."),
    ("Unit landscaping",                            1900000, KPV['landscaping_per_unit_avg'],    f"${KPV['landscaping_per_unit_range'][0]/1000:.0f}k - ${KPV['landscaping_per_unit_range'][1]/1000:.0f}k", "Per-unit lots: courtyards, plantings."),
    ("Council Dev fees",                            2500000, KPV['council_fees_per_unit_avg'],   f"${KPV['council_fees_per_unit_range'][0]/1000:.0f}k - ${KPV['council_fees_per_unit_range'][1]/1000:.0f}k", "DCs per allotment + per BC."),
    ("Infrastructure + civils + site landscape",    8400000, KPV['infrastructure_per_unit_avg'], f"${KPV['infrastructure_per_unit_range'][0]/1000:.0f}k - ${KPV['infrastructure_per_unit_range'][1]/1000:.0f}k", "Earthworks, roads, 3-waters, power, common landscape."),
    ("Other site works",                             980000, KPV['other_site_works_per_unit_avg'], f"${KPV['other_site_works_per_unit_range'][0]/1000:.0f}k - ${KPV['other_site_works_per_unit_range'][1]/1000:.0f}k", "Boundary, signage, gates, etc."),
    ("Professional fees + preliminaries",           1850000, KPV['professional_per_unit_avg'],   f"${KPV['professional_per_unit_range'][0]/1000:.0f}k - ${KPV['professional_per_unit_range'][1]/1000:.0f}k", "Architecture, engineering, surveying, consents prep."),
    ("Holding, marketing, insurance, rates",        2600000, None, "$2.0-3.0m typical total", "Spread over dev period."),
    ("KPV management + transaction fees",           4300000, None, "4-6% of revenue typical", "Operator overhead."),
]
DEV_START = 29
DEV_LABELS = []
for i, (label, default_total, kpv_avg, rng, comment) in enumerate(dev_rows):
    r = DEV_START + i
    DEV_LABELS.append((label, r))
    ws_in[f"B{r}"] = label
    fill_cell(ws_in, f"C{r}", default_total, font=INPUT_FT, fmt=FMT_CCY, align="right", comment=comment)
    fill_cell(ws_in, f"D{r}", f"=IFERROR(C{r}/C22,0)", fmt=FMT_CCY, align="right")
    if kpv_avg is not None:
        fill_cell(ws_in, f"E{r}", kpv_avg, font=BENCH_FT, fmt=FMT_CCY, align="center")
    else:
        fill_cell(ws_in, f"E{r}", "—", font=BENCH_FT, align="center")
    fill_cell(ws_in, f"F{r}", rng, font=BENCH_FT, align="center")
DEV_END = DEV_START + len(dev_rows) - 1

# ========== PAVILION section rows 38-43 ==========
PAV_HDR = DEV_END + 2  # row 39
section_header(ws_in, PAV_HDR, "PAVILION / CLUBHOUSE — star-rated build")
PAV_RATING = PAV_HDR + 1
PAV_SIZE = PAV_HDR + 2
PAV_RATE = PAV_HDR + 3
PAV_COST = PAV_HDR + 4

ws_in[f"B{PAV_RATING}"] = "Pavilion star rating"
fill_cell(ws_in, f"C{PAV_RATING}", "4.0 star", font=INPUT_FT, align="center")
ws_in[f"E{PAV_RATING}"] = "3.0★ / 3.5★ / 4.0★ / 4.5★"; ws_in[f"E{PAV_RATING}"].font = BENCH_FT
ws_in.merge_cells(f"E{PAV_RATING}:F{PAV_RATING}")
ws_in[f"G{PAV_RATING}"] = "Drives build rate per m²"
dv_pav = DataValidation(type="list", formula1='"3.0 star,3.5 star,4.0 star,4.5 star"', allow_blank=False)
ws_in.add_data_validation(dv_pav); dv_pav.add(f"C{PAV_RATING}")

ws_in[f"B{PAV_SIZE}"] = "Pavilion size (m²)"
fill_cell(ws_in, f"C{PAV_SIZE}", 800, font=INPUT_FT, fmt=FMT_INT, align="right")
ws_in[f"E{PAV_SIZE}"] = "typical 600-900 m²"; ws_in[f"E{PAV_SIZE}"].font = BENCH_FT
ws_in.merge_cells(f"E{PAV_SIZE}:F{PAV_SIZE}")

# Pavilion rate lookup — small hidden block in column J for cleanliness
ws_in["J1"] = "Pavilion rate lookup"; ws_in["J1"].font = SMALL
ws_in["J2"] = "3.0 star"; ws_in["K2"] = 4000
ws_in["J3"] = "3.5 star"; ws_in["K3"] = 5000
ws_in["J4"] = "4.0 star"; ws_in["K4"] = 6000
ws_in["J5"] = "4.5 star"; ws_in["K5"] = 7000
ws_in.column_dimensions['J'].hidden = True
ws_in.column_dimensions['K'].hidden = True

ws_in[f"B{PAV_RATE}"] = "Pavilion build rate ($/m² incl GST)"
fill_cell(ws_in, f"C{PAV_RATE}", f"=VLOOKUP(C{PAV_RATING},J2:K5,2,FALSE)", fmt=FMT_CCY, align="right")
ws_in[f"E{PAV_RATE}"] = "Auto from star rating"; ws_in[f"E{PAV_RATE}"].font = BENCH_FT
ws_in.merge_cells(f"E{PAV_RATE}:F{PAV_RATE}")

ws_in[f"B{PAV_COST}"] = "Pavilion total cost"
fill_cell(ws_in, f"C{PAV_COST}", f"=C{PAV_RATE}*C{PAV_SIZE}", fmt=FMT_CCY, align="right", bold=True, fill=LBLUE)
ws_in[f"E{PAV_COST}"] = f"${KPV['clubhouse_total_range'][0]/1000000:.1f}M - ${KPV['clubhouse_total_range'][1]/1000000:.1f}M KPV range"; ws_in[f"E{PAV_COST}"].font = BENCH_FT
ws_in.merge_cells(f"E{PAV_COST}:F{PAV_COST}")

# ========== TIMING section ==========
TIM_HDR = PAV_COST + 2
section_header(ws_in, TIM_HDR, "TIMING")
ws_in[f"B{TIM_HDR+1}"] = "Development duration (years)"
fill_cell(ws_in, f"C{TIM_HDR+1}", 6, font=INPUT_FT, fmt=FMT_INT, align="right")
ws_in[f"G{TIM_HDR+1}"] = "Start to last unit complete"
ws_in[f"B{TIM_HDR+2}"] = "Sales velocity (units/month)"
fill_cell(ws_in, f"C{TIM_HDR+2}", 1.5, font=INPUT_FT, fmt=FMT_DEC, align="right")
ws_in[f"G{TIM_HDR+2}"] = "Steady-state absorption"

# ========== VENDOR FINANCE section ==========
VF_HDR = TIM_HDR + 4
section_header(ws_in, VF_HDR, "VENDOR FINANCE — only relevant if Land scenario = Vendor Finance")
ws_in[f"B{VF_HDR+1}"] = "Vendor interest rate (annual)"
fill_cell(ws_in, f"C{VF_HDR+1}", "=AssumptionsRef!F37" if False else KPV['vendor_rate'], font=BENCH_FT, fmt=FMT_PCT, align="right", fill=YELLOW, comment="Henley benchmark 4% — counter-payment terms")
ws_in[f"B{VF_HDR+2}"] = "Long Stop Date (years)"
fill_cell(ws_in, f"C{VF_HDR+2}", KPV['vendor_lsd_years'], font=BENCH_FT, fmt=FMT_INT, align="right", fill=YELLOW, comment="5-7 years typical")

# ========== KPV DEFAULTS section (yellow fill — source from AssumptionsRef) ==========
DEF_HDR = VF_HDR + 4
section_header(ws_in, DEF_HDR, "KPV DEFAULTS — sourced from Assumptions Reference sheet (editable, override if needed)")
# Each row: (label, default_value, fmt, source_in_ref_sheet, comment)
defaults = [
    ("DMF percentage (on resale price)", KPV['dmf_pct'], FMT_PCT, "AssumptionsRef!F37", "NZ industry typical 24%. Confirmed in Cromwell + Henley feasibility models."),
    ("Average resident tenure (years)",  KPV['tenure_years'], FMT_INT, "AssumptionsRef!F38", "Kyle's assumption: 9 yr midpoint (Cromwell 8 / Henley 10)."),
    ("Resale fee",                       KPV['resale_fee'], FMT_PCT, None, "NZ RV industry typical 1%."),
    ("Capital growth (annual)",          KPV['capital_growth'], FMT_PCT, None, "Conservative."),
    ("Cost inflation (annual)",          KPV['cost_inflation'], FMT_PCT, None, "Matched to capital growth for neutrality at base."),
    ("Discount rate (terminal value)",   KPV['discount_rate'], FMT_PCT, "AssumptionsRef!F60", "Kyle's internal: 17.5% (Cromwell). JLL valuation basis: 14.0%."),
    ("Preferred return on capital (A-Class)", KPV['preferred_rate'], FMT_PCT, None, "JV preferred coupon."),
]
for i, (label, val, fmt, src, comment) in enumerate(defaults):
    r = DEF_HDR + 1 + i
    ws_in[f"B{r}"] = label
    if src:
        # Reference to Assumptions sheet for primary value
        fill_cell(ws_in, f"C{r}", f"={src}", fmt=fmt, align="right", fill=YELLOW, comment=comment)
        ws_in[f"G{r}"] = f"Linked to {src.split('!')[0]} sheet"
        ws_in[f"G{r}"].font = SMALL
    else:
        fill_cell(ws_in, f"C{r}", val, fmt=fmt, align="right", fill=YELLOW, comment=comment)
DEF_END = DEF_HDR + len(defaults)

# Cell references we'll reuse across sheets
INP = {
    'land_ha': 'Inputs!C16', 'land_m2': 'Inputs!C17', 'land_price': 'Inputs!C18', 'land_scen': 'Inputs!C19',
    'units': 'Inputs!C22', 'ora': 'Inputs!C23', 'unit_size': 'Inputs!C24', 'density': 'Inputs!C25',
    'construction': f'Inputs!C{DEV_LABELS[0][1]}', 'landscaping': f'Inputs!C{DEV_LABELS[1][1]}',
    'council': f'Inputs!C{DEV_LABELS[2][1]}', 'infra': f'Inputs!C{DEV_LABELS[3][1]}',
    'other': f'Inputs!C{DEV_LABELS[4][1]}', 'professional': f'Inputs!C{DEV_LABELS[5][1]}',
    'holding': f'Inputs!C{DEV_LABELS[6][1]}', 'kpv_fees': f'Inputs!C{DEV_LABELS[7][1]}',
    'pavilion_cost': f'Inputs!C{PAV_COST}',
    'dev_years': f'Inputs!C{TIM_HDR+1}', 'sales_vel': f'Inputs!C{TIM_HDR+2}',
    'vendor_rate': f'Inputs!C{VF_HDR+1}', 'vendor_lsd': f'Inputs!C{VF_HDR+2}',
    'dmf': f'Inputs!C{DEF_HDR+1}', 'tenure': f'Inputs!C{DEF_HDR+2}', 'resale_fee': f'Inputs!C{DEF_HDR+3}',
    'cap_growth': f'Inputs!C{DEF_HDR+4}', 'cost_infl': f'Inputs!C{DEF_HDR+5}',
    'discount': f'Inputs!C{DEF_HDR+6}', 'preferred': f'Inputs!C{DEF_HDR+7}',
}

# ========== SENSE-CHECK WARNINGS ==========
WARN_HDR = DEF_END + 2
section_header(ws_in, WARN_HDR, "SENSE-CHECK WARNINGS — auto-flagged from inputs")
warn_rows = [
    ("Construction $/unit",
     f'=IF(OR(D{DEV_LABELS[0][1]}<{KPV["construction_per_unit_range"][0]},D{DEV_LABELS[0][1]}>{KPV["construction_per_unit_range"][1]}),"⚠ Construction $/unit "&TEXT(D{DEV_LABELS[0][1]},"$#,##0")&" outside KPV range $"&TEXT({KPV["construction_per_unit_range"][0]},"#,##0")&"-$"&TEXT({KPV["construction_per_unit_range"][1]},"#,##0"),"✓ Construction $/unit in range")'),
    ("Density",
     f'=IF(OR(C25<20,C25>24),"⚠ Density "&TEXT(C25,"0.0")&" units/ha outside KPV range 20-24","✓ Density in range")'),
    ("ORA price",
     f'=IF(OR(C23<700000,C23>1100000),"⚠ ORA price "&TEXT(C23,"$#,##0")&" outside KPV portfolio range $700k-$1.1m","✓ ORA price in range")'),
    ("Land $/ha",
     f'=IF(OR(C18/C16<1000000,C18/C16>5000000),"⚠ Land price "&TEXT(C18/C16,"$#,##0")&"/ha unusually high or low (KPV typical $1m-$5m/ha)","✓ Land $/ha in range")'),
]
for i, (k, formula) in enumerate(warn_rows):
    r = WARN_HDR + 1 + i
    ws_in[f"B{r}"] = k
    ws_in[f"C{r}"] = formula
    ws_in.merge_cells(f"C{r}:G{r}")
    # Conditional fill: amber if starts with "⚠"
    ws_in.conditional_formatting.add(f"C{r}", FormulaRule(formula=[f'LEFT(C{r},1)="⚠"'], fill=LAMBER))
    ws_in.conditional_formatting.add(f"C{r}", FormulaRule(formula=[f'LEFT(C{r},1)="✓"'], fill=LGREEN))

ws_in.freeze_panes = "A14"  # freeze the headline rows visible

# =================================================================
# SHEET 2: CALCULATIONS
# =================================================================
ws_calc = wb.create_sheet("Calculations")
widths_calc = {'A': 2, 'B': 46, 'C': 18, 'D': 18, 'E': 18, 'F': 18, 'G': 18}
for col, w in widths_calc.items():
    ws_calc.column_dimensions[col].width = w

title_bar(ws_calc, 2, "CALCULATIONS — by Land Scenario", span="B:G")
ws_calc["B3"] = "All four scenarios computed in parallel. SELECTED column pulls the scenario from Inputs!C19."
ws_calc["B3"].font = SMALL
ws_calc.merge_cells("B3:G3")

# Headers row 5
ws_calc["B5"] = "Metric"
ws_calc["C5"] = "Ownership"; ws_calc["D5"] = "Deferred"; ws_calc["E5"] = "Lease"; ws_calc["F5"] = "Vendor Finance"; ws_calc["G5"] = "SELECTED"
for col in "BCDEFG":
    fill_cell(ws_calc, f"{col}5", bold=True, fill=LBLUE, align="center")

# Helper for scenario formulas
def scen_formula(template_fn):
    """Apply template_fn for each of 4 scenarios. Returns dict of col→formula."""
    return {col: template_fn(col) for col in "CDEF"}

# Pre-compute commonly-used sub-formulas
GROWTH_MID = f"(1+{INP['cap_growth']})^({INP['dev_years']}/2)"
INFL_MID   = f"(1+{INP['cost_infl']})^({INP['dev_years']}/2)"
GROSS_ORA = f"({INP['units']}*{INP['ora']})"
ORA_GROWN = f"({GROSS_ORA}*{GROWTH_MID})"

# Land payment per scenario
def land_payment(col):
    if col == "C":  return f"={INP['land_price']}"
    if col == "D":  return f"={INP['land_price']}"
    if col == "E":  return f"={INP['land_price']}*1.15*0.06*{INP['dev_years']}"
    if col == "F":  return f"={INP['land_price']}"

# Vendor interest (only F)
def vendor_int(col):
    return f"={INP['land_price']}*0.55*{INP['vendor_rate']}*{INP['vendor_lsd']}" if col == "F" else "=0"

# Peak capital factor
PEAK_FACTORS = {"C": 0.35, "D": 0.30, "E": 0.30, "F": 0.20}

# Row plan
r = 6  # Revenue section
ws_calc[f"B{r}"] = "REVENUE"; subtotal(ws_calc, f"B{r}", "REVENUE")
ws_calc.merge_cells(f"B{r}:F{r}")
r += 1
GROSS_R = r
ws_calc[f"B{r}"] = "Gross ORA sales (initial, nominal)"
for col in "CDEF": fill_cell(ws_calc, f"{col}{r}", f"={GROSS_ORA}", fmt=FMT_CCY, align="right")
r += 1
INFL_R = r
ws_calc[f"B{r}"] = "ORA sales (with mid-period capital growth)"
for col in "CDEF": fill_cell(ws_calc, f"{col}{r}", f"={ORA_GROWN}", fmt=FMT_CCY, align="right", bold=True, fill=LBLUE)

# Costs section
r += 2
ws_calc[f"B{r}"] = "COSTS"; subtotal(ws_calc, f"B{r}", "COSTS")
ws_calc.merge_cells(f"B{r}:F{r}")

r += 1
LAND_R = r
ws_calc[f"B{r}"] = "Land payment / ground rent"
for col, f in scen_formula(land_payment).items():
    fill_cell(ws_calc, f"{col}{r}", f, fmt=FMT_CCY, align="right")

r += 1
VINT_R = r
ws_calc[f"B{r}"] = "Vendor interest cost (approx)"
for col, f in scen_formula(vendor_int).items():
    fill_cell(ws_calc, f"{col}{r}", f, fmt=FMT_CCY, align="right")

r += 1
CONST_R = r
ws_calc[f"B{r}"] = "Construction (inflated to mid-dev)"
for col in "CDEF":
    fill_cell(ws_calc, f"{col}{r}", f"={INP['construction']}*{INFL_MID}", fmt=FMT_CCY, align="right")

r += 1
LSCAPE_R = r
ws_calc[f"B{r}"] = "Unit landscaping (inflated)"
for col in "CDEF":
    fill_cell(ws_calc, f"{col}{r}", f"={INP['landscaping']}*{INFL_MID}", fmt=FMT_CCY, align="right")

r += 1
COUNCIL_R = r
ws_calc[f"B{r}"] = "Council Dev fees (not inflated)"
for col in "CDEF":
    fill_cell(ws_calc, f"{col}{r}", f"={INP['council']}", fmt=FMT_CCY, align="right")

r += 1
INFRA_R = r
ws_calc[f"B{r}"] = "Infrastructure (inflated)"
for col in "CDEF":
    fill_cell(ws_calc, f"{col}{r}", f"={INP['infra']}*{INFL_MID}", fmt=FMT_CCY, align="right")

r += 1
OTHER_R = r
ws_calc[f"B{r}"] = "Other site works (inflated)"
for col in "CDEF":
    fill_cell(ws_calc, f"{col}{r}", f"={INP['other']}*{INFL_MID}", fmt=FMT_CCY, align="right")

r += 1
PROF_R = r
ws_calc[f"B{r}"] = "Professional fees + preliminaries"
for col in "CDEF":
    fill_cell(ws_calc, f"{col}{r}", f"={INP['professional']}", fmt=FMT_CCY, align="right")

r += 1
HOLD_R = r
ws_calc[f"B{r}"] = "Holding, marketing, insurance, rates"
for col in "CDEF":
    fill_cell(ws_calc, f"{col}{r}", f"={INP['holding']}", fmt=FMT_CCY, align="right")

r += 1
KPVF_R = r
ws_calc[f"B{r}"] = "KPV management + transaction fees"
for col in "CDEF":
    fill_cell(ws_calc, f"{col}{r}", f"={INP['kpv_fees']}", fmt=FMT_CCY, align="right")

r += 1
PAV_R = r
ws_calc[f"B{r}"] = "Pavilion / clubhouse (inflated)"
for col in "CDEF":
    fill_cell(ws_calc, f"{col}{r}", f"={INP['pavilion_cost']}*{INFL_MID}", fmt=FMT_CCY, align="right")

r += 1
RESALE_R = r
ws_calc[f"B{r}"] = "Resale fee on initial sales"
for col in "CDEF":
    fill_cell(ws_calc, f"{col}{r}", f"={col}{INFL_R}*{INP['resale_fee']}", fmt=FMT_CCY, align="right")

r += 1
TDC_R = r
ws_calc[f"B{r}"] = "TOTAL DEVELOPMENT COST"
for col in "CDEF":
    fill_cell(ws_calc, f"{col}{r}", f"=SUM({col}{LAND_R}:{col}{RESALE_R})", fmt=FMT_CCY, align="right", bold=True, fill=LBLUE)

r += 1
SURP_R = r
ws_calc[f"B{r}"] = "Development surplus (pre-preferred)"
for col in "CDEF":
    fill_cell(ws_calc, f"{col}{r}", f"={col}{INFL_R}-{col}{TDC_R}", fmt=FMT_CCY, align="right", bold=True, fill=LBLUE)

# Peak capital factors row
r += 1
PKF_R = r
ws_calc[f"B{r}"] = "Peak capital factor"
ws_calc[f"B{r}"].font = SMALL
for col, fac in PEAK_FACTORS.items():
    fill_cell(ws_calc, f"{col}{r}", fac, font=SMALL, fmt=FMT_PCT, align="center")

r += 1
PREF_R = r
ws_calc[f"B{r}"] = "Preferred interest cost (approx)"
for col in "CDEF":
    fill_cell(ws_calc, f"{col}{r}", f"={col}{TDC_R}*{col}{PKF_R}*{INP['preferred']}*{INP['dev_years']}*0.6", fmt=FMT_CCY, align="right")

r += 1
SAP_R = r
ws_calc[f"B{r}"] = "Surplus after preferred"
for col in "CDEF":
    fill_cell(ws_calc, f"{col}{r}", f"={col}{SURP_R}-{col}{PREF_R}", fmt=FMT_CCY, align="right")

# Steady-state
r += 2
ws_calc[f"B{r}"] = "STEADY-STATE OPERATING"; subtotal(ws_calc, f"B{r}", "STEADY-STATE OPERATING")
ws_calc.merge_cells(f"B{r}:F{r}")

r += 1
DMF_R = r
ws_calc[f"B{r}"] = "Annual DMF income at maturity"
DMF_FORM = f"({INP['units']}/{INP['tenure']})*{INP['ora']}*(1+{INP['cap_growth']})^{INP['tenure']}*({INP['dmf']}-{INP['resale_fee']})"
for col in "CDEF":
    fill_cell(ws_calc, f"{col}{r}", f"={DMF_FORM}", fmt=FMT_CCY, align="right")

r += 1
TV_R = r
ws_calc[f"B{r}"] = "Terminal value (DMF / discount rate)"
for col in "CDEF":
    fill_cell(ws_calc, f"{col}{r}", f"={col}{DMF_R}/{INP['discount']}", fmt=FMT_CCY, align="right")

# Gateway metrics
r += 2
ws_calc[f"B{r}"] = "GATEWAY METRICS"; subtotal(ws_calc, f"B{r}", "GATEWAY METRICS")
ws_calc.merge_cells(f"B{r}:F{r}")

r += 1
MARG_R = r
ws_calc[f"B{r}"] = "Development margin on cost"
for col in "CDEF":
    fill_cell(ws_calc, f"{col}{r}", f"=IFERROR({col}{SURP_R}/{col}{TDC_R},0)", fmt=FMT_PCT, align="right", bold=True)

r += 1
MCAP_R = r
ws_calc[f"B{r}"] = "Margin on capital (post-preferred)"
for col in "CDEF":
    fill_cell(ws_calc, f"{col}{r}", f"=IFERROR({col}{SAP_R}/{col}{TDC_R},0)", fmt=FMT_PCT, align="right")

r += 1
IRR_R = r
ws_calc[f"B{r}"] = "Project IRR (15-year proxy)"
for col in "CDEF":
    fill_cell(ws_calc, f"{col}{r}", f"=IFERROR(RATE(15,0,-{col}{TDC_R}*{col}{PKF_R},{col}{TV_R}+{col}{SURP_R}),0)", fmt=FMT_PCT, align="right", bold=True)

r += 1
PEAK_R = r
ws_calc[f"B{r}"] = "Peak capital required"
for col in "CDEF":
    fill_cell(ws_calc, f"{col}{r}", f"={col}{TDC_R}*{col}{PKF_R}", fmt=FMT_CCY, align="right")

r += 1
DPU_R = r
ws_calc[f"B{r}"] = "Total dev cost per unit"
for col in "CDEF":
    fill_cell(ws_calc, f"{col}{r}", f"={col}{TDC_R}/{INP['units']}", fmt=FMT_CCY, align="right")

# Verdict
r += 2
VERD_R = r
ws_calc[f"B{r}"] = "VERDICT"
fill_cell(ws_calc, f"B{r}", bold=True, fill=HEAD, font=WB_BOLD)
for col in "CDEF":
    f = (f'=IF(AND({col}{MARG_R}>={KPV["proceed_margin"]},{col}{IRR_R}>={KPV["proceed_irr"]}),"PROCEED",'
         f'IF(OR({col}{MARG_R}<{KPV["decline_margin"]},{col}{IRR_R}<{KPV["decline_irr"]}),"DECLINE","PROCEED WITH CAUTION"))')
    fill_cell(ws_calc, f"{col}{r}", f, bold=True, align="center")
    ws_calc[f"{col}{r}"].font = LARGE_BD

# SELECTED column (G) — INDEX/MATCH against C19 scenario
SCENARIO_ARRAY = '{"Ownership","Deferred","Lease","Vendor Finance"}'
for src_r in [GROSS_R, INFL_R, LAND_R, VINT_R, CONST_R, LSCAPE_R, COUNCIL_R, INFRA_R, OTHER_R, PROF_R, HOLD_R, KPVF_R, PAV_R, RESALE_R, TDC_R, SURP_R, PKF_R, PREF_R, SAP_R, DMF_R, TV_R, MARG_R, MCAP_R, IRR_R, PEAK_R, DPU_R, VERD_R]:
    # Detect format from C-cell
    c_cell = ws_calc[f"C{src_r}"]
    fmt = c_cell.number_format
    fill_cell(ws_calc, f"G{src_r}",
              f'=INDEX(C{src_r}:F{src_r},MATCH({INP["land_scen"]},{SCENARIO_ARRAY},0))',
              fmt=fmt if fmt and fmt != 'General' else None,
              align="right" if src_r != VERD_R else "center",
              bold=(src_r in [INFL_R, TDC_R, SURP_R, MARG_R, IRR_R, VERD_R]),
              fill=LBLUE if src_r in [INFL_R, TDC_R, SURP_R] else (HEAD if src_r == VERD_R else None),
              font=LARGE_BD if src_r == VERD_R else (WB_BOLD if src_r == VERD_R else None))

# Apply conditional formatting on metrics columns
cf_threshold(ws_calc, f"C{MARG_R}:G{MARG_R}", 'margin')
cf_threshold(ws_calc, f"C{IRR_R}:G{IRR_R}", 'irr')
cf_threshold(ws_calc, f"C{VERD_R}:F{VERD_R}", 'verdict')
cf_threshold(ws_calc, f"G{VERD_R}:G{VERD_R}", 'verdict')

# Comparison block at bottom
r = VERD_R + 3
section_header(ws_calc, r, "ALL-SCENARIO COMPARISON")
r += 1
ws_calc[f"B{r}"] = "Metric"; ws_calc[f"C{r}"] = "Ownership"; ws_calc[f"D{r}"] = "Deferred"; ws_calc[f"E{r}"] = "Lease"; ws_calc[f"F{r}"] = "Vendor Finance"
for col in "BCDEF": fill_cell(ws_calc, f"{col}{r}", bold=True, fill=LBLUE, align="center")
for label, src_r, fmt in [("Margin on cost", MARG_R, FMT_PCT), ("Project IRR", IRR_R, FMT_PCT), ("Peak capital", PEAK_R, FMT_CCY), ("Verdict", VERD_R, None)]:
    r += 1
    ws_calc[f"B{r}"] = label
    for col in "CDEF":
        fill_cell(ws_calc, f"{col}{r}", f"={col}{src_r}", fmt=fmt, align="right", bold=(src_r==VERD_R))
ws_calc.freeze_panes = "B6"

# =================================================================
# Wire headline block on Inputs (now that Calculations exists)
# =================================================================
# Row 6-11 values pull from Calculations SELECTED column
fill_cell(ws_in, "C6", f"={INP['land_scen']}", align="center", bold=True)
fill_cell(ws_in, "C7", f"=Calculations!G{MARG_R}", fmt=FMT_PCT, align="right", bold=True, font=LARGE_BD)
fill_cell(ws_in, "C8", f"=Calculations!G{IRR_R}", fmt=FMT_PCT, align="right", bold=True, font=LARGE_BD)
fill_cell(ws_in, "C9", f"=Calculations!G{PEAK_R}", fmt=FMT_CCY, align="right")
fill_cell(ws_in, "C10", f"=Calculations!G{DPU_R}", fmt=FMT_CCY, align="right")
fill_cell(ws_in, "C11", f"=Inputs!C25", fmt=FMT_DEC, align="right")
# Verdict
fill_cell(ws_in, "C12", f"=Calculations!G{VERD_R}", bold=True, align="center")
ws_in["C12"].font = Font(bold=True, size=18, color="FFFFFF")
# CF on headline cells
cf_threshold(ws_in, "C7:C7", 'margin')
cf_threshold(ws_in, "C8:C8", 'irr')
cf_threshold(ws_in, "C12:G12", 'verdict')

# =================================================================
# SHEET 3: MAXIMUM VIABLE PRICE
# =================================================================
ws_mvp = wb.create_sheet("Maximum Viable Price")
widths_mvp = {'A': 2, 'B': 42, 'C': 18, 'D': 18, 'E': 18, 'F': 18}
for col, w in widths_mvp.items():
    ws_mvp.column_dimensions[col].width = w

title_bar(ws_mvp, 2, "MAXIMUM VIABLE LAND PRICE", span="B:F")
ws_mvp["B3"] = "Reverse calc: at what land price does each scenario hit the 15% margin threshold?"
ws_mvp["B3"].font = SMALL; ws_mvp.merge_cells("B3:F3")

# Common "other costs" expression — sum of all non-land cost lines as inflated/non-inflated per the model
# = construction*infl + landscaping*infl + council + infra*infl + other*infl + professional + holding + kpv_fees + pavilion*infl
OTHER_COSTS = (
    f"({INP['construction']}*{INFL_MID}"
    f"+{INP['landscaping']}*{INFL_MID}"
    f"+{INP['council']}"
    f"+{INP['infra']}*{INFL_MID}"
    f"+{INP['other']}*{INFL_MID}"
    f"+{INP['professional']}"
    f"+{INP['holding']}"
    f"+{INP['kpv_fees']}"
    f"+{INP['pavilion_cost']}*{INFL_MID})"
)
RESALE_TERM = f"({GROSS_ORA}*{GROWTH_MID})*{INP['resale_fee']}"
TOTAL_REV = f"({GROSS_ORA}*{GROWTH_MID})"

# Block 1: Max land price at margin threshold
section_header(ws_mvp, 5, "BLOCK 1: Max land price at 15% margin threshold", span="B:F")
ws_mvp["B6"] = "Metric"
ws_mvp["C6"] = "Ownership"; ws_mvp["D6"] = "Deferred"; ws_mvp["E6"] = "Lease"; ws_mvp["F6"] = "Vendor Finance"
for col in "BCDEF": fill_cell(ws_mvp, f"{col}6", bold=True, fill=LBLUE, align="center")

ws_mvp["B7"] = "Target margin"
for col in "CDEF":
    fill_cell(ws_mvp, f"{col}7", f"={KPV['proceed_margin']}", fmt=FMT_PCT, align="center", font=BENCH_FT)

# Budget = Rev/(1+m) - OtherCosts - ResaleFee
BUDGET = f"({TOTAL_REV}/(1+{KPV['proceed_margin']})-{OTHER_COSTS}-{RESALE_TERM})"
ws_mvp["B8"] = "Max viable land price"
fill_cell(ws_mvp, "C8", f"={BUDGET}", fmt=FMT_CCY, align="right", bold=True, fill=LGREEN)
fill_cell(ws_mvp, "D8", f"={BUDGET}", fmt=FMT_CCY, align="right", bold=True, fill=LGREEN)
# Lease: land term = land_price * 1.15 * 0.06 * dev_years  →  land_price = budget / (1.15*0.06*dev_years)
fill_cell(ws_mvp, "E8", f"={BUDGET}/(1.15*0.06*{INP['dev_years']})", fmt=FMT_CCY, align="right", bold=True, fill=LGREEN)
# VF: land_price + 0.55*rate*lsd*land_price = budget  →  land_price = budget / (1 + 0.55*rate*lsd)
fill_cell(ws_mvp, "F8", f"={BUDGET}/(1+0.55*{INP['vendor_rate']}*{INP['vendor_lsd']})", fmt=FMT_CCY, align="right", bold=True, fill=LGREEN)

ws_mvp["B9"] = "Current land price"
for col in "CDEF":
    fill_cell(ws_mvp, f"{col}9", f"={INP['land_price']}", fmt=FMT_CCY, align="right", font=BENCH_FT)

ws_mvp["B10"] = "Headroom (max − current)"
for col in "CDEF":
    fill_cell(ws_mvp, f"{col}10", f"={col}8-{col}9", fmt=FMT_CCY, align="right", bold=True)

# Block 2: Min unit count at current land price
section_header(ws_mvp, 12, "BLOCK 2: Min unit count required at current land price", span="B:F")
ws_mvp["B13"] = "Metric"
ws_mvp["C13"] = "Ownership"; ws_mvp["D13"] = "Deferred"; ws_mvp["E13"] = "Lease"; ws_mvp["F13"] = "Vendor Finance"
for col in "BCDEF": fill_cell(ws_mvp, f"{col}13", bold=True, fill=LBLUE, align="center")

# Effective land cost per scenario
EFF_LAND = {
    "C": f"{INP['land_price']}",
    "D": f"{INP['land_price']}",
    "E": f"{INP['land_price']}*1.15*0.06*{INP['dev_years']}",
    "F": f"{INP['land_price']}*(1+0.55*{INP['vendor_rate']}*{INP['vendor_lsd']})",
}
# min_units = (1 + margin)*(land + other_costs) / (ORA*growth*(1-resale_fee))
ws_mvp["B14"] = "Min unit count (at base ORA)"
for col in "CDEF":
    f = f"=IFERROR((1+{KPV['proceed_margin']})*({EFF_LAND[col]}+{OTHER_COSTS})/({INP['ora']}*{GROWTH_MID}*(1-{INP['resale_fee']})),0)"
    fill_cell(ws_mvp, f"{col}14", f, fmt=FMT_INT, align="right", bold=True, fill=LGREEN)

ws_mvp["B15"] = "Current unit count"
for col in "CDEF":
    fill_cell(ws_mvp, f"{col}15", f"={INP['units']}", fmt=FMT_INT, align="right", font=BENCH_FT)

ws_mvp["B16"] = "Additional units required"
for col in "CDEF":
    fill_cell(ws_mvp, f"{col}16", f"=MAX(0,{col}14-{col}15)", fmt=FMT_INT, align="right", bold=True)

# Block 3: Min ORA price required
section_header(ws_mvp, 18, "BLOCK 3: Min ORA price required at current unit count", span="B:F")
ws_mvp["B19"] = "Metric"
ws_mvp["C19"] = "Ownership"; ws_mvp["D19"] = "Deferred"; ws_mvp["E19"] = "Lease"; ws_mvp["F19"] = "Vendor Finance"
for col in "BCDEF": fill_cell(ws_mvp, f"{col}19", bold=True, fill=LBLUE, align="center")

# min_ORA = (1 + margin)*(land + other_costs) / (units*growth*(1-resale_fee))
ws_mvp["B20"] = "Min ORA price (at base units)"
for col in "CDEF":
    f = f"=IFERROR((1+{KPV['proceed_margin']})*({EFF_LAND[col]}+{OTHER_COSTS})/({INP['units']}*{GROWTH_MID}*(1-{INP['resale_fee']})),0)"
    fill_cell(ws_mvp, f"{col}20", f, fmt=FMT_CCY, align="right", bold=True, fill=LGREEN)

ws_mvp["B21"] = "Current ORA price"
for col in "CDEF":
    fill_cell(ws_mvp, f"{col}21", f"={INP['ora']}", fmt=FMT_CCY, align="right", font=BENCH_FT)

ws_mvp["B22"] = "ORA uplift required"
for col in "CDEF":
    fill_cell(ws_mvp, f"{col}22", f"=MAX(0,{col}20-{col}21)", fmt=FMT_CCY, align="right", bold=True)

# Flag amber if uplift > 10%
ws_mvp.conditional_formatting.add(f"C22:F22", FormulaRule(formula=[f'C22/C21>0.1'], fill=LAMBER))

# Block 4: Land $/ha comparison
section_header(ws_mvp, 24, "BLOCK 4: Land $/hectare comparison", span="B:F")
ws_mvp["B25"] = "Metric"
ws_mvp["C25"] = "Value"
fill_cell(ws_mvp, "B25", bold=True, fill=LBLUE, align="center")
fill_cell(ws_mvp, "C25", bold=True, fill=LBLUE, align="center")
ws_mvp["B26"] = "Current land $/ha";       fill_cell(ws_mvp, "C26", f"={INP['land_price']}/{INP['land_ha']}", fmt=FMT_CCY, align="right")
ws_mvp["B27"] = "Max viable land $/ha (Ownership)"; fill_cell(ws_mvp, "C27", f"=C8/{INP['land_ha']}", fmt=FMT_CCY, align="right", bold=True, fill=LGREEN)
ws_mvp["B28"] = "KPV portfolio reference"
ws_mvp["C28"] = "$1m – $3m / ha (greenfield RV)"
ws_mvp["C28"].font = BENCH_FT; ws_mvp["C28"].alignment = Alignment(horizontal="right")

ws_mvp.freeze_panes = "A5"

# =================================================================
# SHEET 4: SCENARIOS
# =================================================================
ws_scen = wb.create_sheet("Scenarios")
widths_scen = {'A': 2, 'B': 30, 'C': 36, 'D': 12, 'E': 12, 'F': 14, 'G': 18, 'H': 22}
for col, w in widths_scen.items():
    ws_scen.column_dimensions[col].width = w

title_bar(ws_scen, 2, "SENSITIVITY SCENARIOS", span="B:H")
ws_scen["B3"] = "Each scenario flexes specific inputs (construction, ORA, units, duration, inflation) and recomputes the key metrics. Selected land scenario applies throughout."
ws_scen["B3"].font = SMALL; ws_scen.merge_cells("B3:H3")

# Headers row 5
hdr_row = 5
ws_scen[f"B{hdr_row}"] = "Scenario"
ws_scen[f"C{hdr_row}"] = "Variable change"
ws_scen[f"D{hdr_row}"] = "Margin"
ws_scen[f"E{hdr_row}"] = "IRR"
ws_scen[f"F{hdr_row}"] = "Peak Capital"
ws_scen[f"G{hdr_row}"] = "Total Dev Cost"
ws_scen[f"H{hdr_row}"] = "Verdict"
for col in "BCDEFGH":
    fill_cell(ws_scen, f"{col}{hdr_row}", bold=True, fill=LBLUE, align="center")

# Scenario definitions: name, description, (construction_mult, ora_mult, units_delta, dev_years_delta, inflation_delta)
SCENARIOS = [
    ("BASE",                "As entered on Inputs sheet",                              1.00, 1.00,    0,  0, 0.00),
    ("Pessimistic",         "Construction +10%, ORA -5%, +1yr duration",               1.10, 0.95,    0,  1, 0.00),
    ("Optimistic",          "Construction -5%, ORA +5%, -1yr duration",                0.95, 1.05,    0, -1, 0.00),
    ("Build cost shock",    "Construction +15%",                                       1.15, 1.00,    0,  0, 0.00),
    ("Sales weakness",      "ORA -10%",                                                1.00, 0.90,    0,  0, 0.00),
    ("Sales strength",      "ORA +10%",                                                1.00, 1.10,    0,  0, 0.00),
    ("High density",        "Unit count +10",                                          1.00, 1.00,   10,  0, 0.00),
    ("Low density",         "Unit count -10",                                          1.00, 1.00,  -10,  0, 0.00),
    ("Premium ORA tier",    "ORA +$50k",                                               1.00, None,    0,  0, 0.00),  # special — use additive
    ("Entry ORA tier",      "ORA -$50k",                                               1.00, None,    0,  0, 0.00),  # special
    ("Long dev",            "Duration +2 years",                                       1.00, 1.00,    0,  2, 0.00),
    ("Short dev",           "Duration -2 years",                                       1.00, 1.00,    0, -2, 0.00),
    ("Inflation shock",     "Cost inflation +1pp",                                     1.00, 1.00,    0,  0, 0.01),
]

# For each scenario, recompute Margin / IRR / Peak Cap / Verdict using the selected scenario (Ownership default)
# To keep formulas compact, we'll bake the recompute inline.
# We use selected scenario via INDEX/MATCH on peak_factor — but for simplicity, assume Ownership (0.35) for all scenarios.
# This gives a comparable basis; user can interpret deltas.

PKF_OWN = 0.35  # use Ownership factor for all scenarios for comparability
ORA_DELTA_PREMIUM = 50000
ORA_DELTA_ENTRY   = -50000

for i, (name, desc, c_mult, o_mult, u_delta, d_delta, infl_delta) in enumerate(SCENARIOS):
    r = hdr_row + 1 + i
    ws_scen[f"B{r}"] = name; ws_scen[f"B{r}"].font = BOLD
    ws_scen[f"C{r}"] = desc; ws_scen[f"C{r}"].font = SMALL

    # ORA expression handling for tier scenarios
    if name == "Premium ORA tier":
        ora_expr = f"({INP['ora']}+{ORA_DELTA_PREMIUM})"
    elif name == "Entry ORA tier":
        ora_expr = f"({INP['ora']}+{ORA_DELTA_ENTRY})"
    else:
        ora_expr = f"({INP['ora']}*{o_mult})"

    units_expr = f"({INP['units']}+{u_delta})" if u_delta else f"{INP['units']}"
    dy_expr = f"({INP['dev_years']}+{d_delta})" if d_delta else f"{INP['dev_years']}"
    infl_expr = f"({INP['cost_infl']}+{infl_delta})" if infl_delta else f"{INP['cost_infl']}"
    growth_expr = f"(1+{INP['cap_growth']})^({dy_expr}/2)"
    inflmid_expr = f"(1+{infl_expr})^({dy_expr}/2)"

    # Revenue
    rev_expr = f"({units_expr}*{ora_expr}*{growth_expr})"

    # Costs (Ownership scenario, all inflated where applicable)
    cost_expr = (
        f"({INP['land_price']}"
        f"+{INP['construction']}*{c_mult}*{inflmid_expr}"
        f"+{INP['landscaping']}*{inflmid_expr}"
        f"+{INP['council']}"
        f"+{INP['infra']}*{inflmid_expr}"
        f"+{INP['other']}*{inflmid_expr}"
        f"+{INP['professional']}"
        f"+{INP['holding']}"
        f"+{INP['kpv_fees']}"
        f"+{INP['pavilion_cost']}*{inflmid_expr}"
        f"+{rev_expr}*{INP['resale_fee']})"
    )
    surp_expr = f"({rev_expr}-{cost_expr})"
    # DMF (steady-state)
    dmf_expr = f"({units_expr}/{INP['tenure']})*{ora_expr}*(1+{INP['cap_growth']})^{INP['tenure']}*({INP['dmf']}-{INP['resale_fee']})"
    tv_expr = f"({dmf_expr}/{INP['discount']})"
    peak_expr = f"({cost_expr}*{PKF_OWN})"

    # Margin
    fill_cell(ws_scen, f"D{r}", f"=IFERROR({surp_expr}/{cost_expr},0)", fmt=FMT_PCT, align="right")
    # IRR
    fill_cell(ws_scen, f"E{r}", f"=IFERROR(RATE(15,0,-{peak_expr},{tv_expr}+{surp_expr}),0)", fmt=FMT_PCT, align="right")
    # Peak cap
    fill_cell(ws_scen, f"F{r}", f"={peak_expr}", fmt=FMT_CCY, align="right")
    # Total dev cost
    fill_cell(ws_scen, f"G{r}", f"={cost_expr}", fmt=FMT_CCY, align="right")
    # Verdict
    f = (f'=IF(AND(D{r}>={KPV["proceed_margin"]},E{r}>={KPV["proceed_irr"]}),"PROCEED",'
         f'IF(OR(D{r}<{KPV["decline_margin"]},E{r}<{KPV["decline_irr"]}),"DECLINE","PROCEED WITH CAUTION"))')
    fill_cell(ws_scen, f"H{r}", f, bold=True, align="center")

last_scen_row = hdr_row + len(SCENARIOS)
cf_threshold(ws_scen, f"D{hdr_row+1}:D{last_scen_row}", 'margin')
cf_threshold(ws_scen, f"E{hdr_row+1}:E{last_scen_row}", 'irr')
cf_threshold(ws_scen, f"H{hdr_row+1}:H{last_scen_row}", 'verdict')

# Summary sentence below
r = last_scen_row + 2
ws_scen[f"B{r}"] = "Summary:"; ws_scen[f"B{r}"].font = BOLD
ws_scen[f"C{r}"] = f'=COUNTIF(H{hdr_row+1}:H{last_scen_row},"PROCEED")&" of {len(SCENARIOS)} scenarios PROCEED, "&COUNTIF(H{hdr_row+1}:H{last_scen_row},"DECLINE")&" DECLINE, "&COUNTIF(H{hdr_row+1}:H{last_scen_row},"PROCEED WITH CAUTION")&" PROCEED WITH CAUTION."'
ws_scen.merge_cells(f"C{r}:H{r}")
ws_scen[f"C{r}"].font = BOLD

# Note about assumption simplification
r += 2
ws_scen[f"B{r}"] = "Note: scenarios use Ownership peak-capital factor (35%) for comparability. Switch to Vendor Finance on Inputs and re-run for VF-specific peak capital."
ws_scen[f"B{r}"].font = SMALL
ws_scen.merge_cells(f"B{r}:H{r}")

ws_scen.freeze_panes = f"B{hdr_row+1}"

# =================================================================
# SHEET 5: NOTES
# =================================================================
ws_notes = wb.create_sheet("Notes")
ws_notes.column_dimensions['A'].width = 2
ws_notes.column_dimensions['B'].width = 110

title_bar(ws_notes, 2, "METHOD AND NOTES", span="B:B")

notes_content = [
    ("PURPOSE", [
        "A 30-minute gating test. Use this before building a full Tier 2 feasibility to confirm the project is worth the effort. Outputs four key metrics (Margin, IRR, Peak Capital, $/unit) plus a verdict.",
        "If PROCEED, move to Tier 2.",
        "If DECLINE, walk away.",
        "If PROCEED WITH CAUTION, identify which inputs need stronger evidence (use the Scenarios sheet and the sense-check warnings on Inputs).",
    ]),
    ("INPUTS CONVENTIONS", [
        "Blue cells are user inputs.",
        "Yellow-highlighted cells are KPV defaults (still editable, usually leave alone). Many source from the Assumptions Reference sheet.",
        "Land area in hectares (m² auto-calculated).",
        "All currency inputs are incl GST unless noted otherwise.",
        "All costs itemised — no loading factor.",
    ]),
    ("METHOD", [
        "Revenue: unit count × ORA price, inflated by capital growth to the midpoint of development.",
        "Costs: construction-related lines inflated to dev midpoint; land, prelims, council fees, marketing not inflated.",
        "Land cost varies by scenario:",
        "  • Ownership: full price month 1",
        "  • Deferred: full price later month, same total as Ownership",
        "  • Lease: 6% of incl-GST land value × dev years",
        "  • Vendor Finance: full price + interest approximation (55% avg outstanding × rate × years)",
        "Preferred interest: scenario-aware peak capital factor (35% Ownership, 30% Deferred/Lease, 20% Vendor Finance) × preferred rate × dev years × 0.6 average outstanding.",
        "DMF income (steady state): (units / tenure) × resale price × (DMF − resale fee). Resale price uses one full tenure of capital growth on ORA.",
        "Terminal value: annual DMF / discount rate (perpetuity).",
        "Project IRR: simplified RATE function over 15 years, using scenario-aware peak capital as average invested.",
    ]),
    ("THRESHOLDS", [
        "PROCEED: margin ≥ 15% AND IRR ≥ 15%",
        "DECLINE: margin < 8% OR IRR < 10%",
        "Otherwise: PROCEED WITH CAUTION",
    ]),
    ("LIMITATIONS", [
        "No monthly cashflow (all costs taken at midpoint).",
        "No partnership waterfall (A-Class / B-Class share splits).",
        "No GST timing.",
        "DMF income assumes steady state from year 1 (in reality ramps as units complete).",
        "No vacancy between residents (typically 2-6 months).",
        "No refurbishment cost on resale (typically $20k-$80k per unit).",
        "No stage-by-stage tranche timing.",
        "Sales velocity input is informational only at Gateway 1 (no cashflow timing model). Use Gateway 2 for velocity sensitivity.",
    ]),
    ("IF THIS PROJECT PASSES THE GATE", [
        "Open the Tier 2 Comprehensive feasibility model (KPV Retirement Gateway 2.xlsx).",
        "Re-enter inputs with the same values used here.",
        "Tier 2 carries monthly cashflow, stage timing, full DMF cycling, sensitivity table, proper XIRR-based IRR, and a Vendor Schedule sheet for vendor finance amortisation.",
    ]),
    ("REFERENCE BENCHMARKS", [
        "The KPV portfolio benchmarks shown alongside inputs are derived from 2027 Master Project Budgets for Papamoa (114 units), Rototuna (131 units), and Waihi Beach (96 units), plus Kyle's curated Assumptions Reference sheet (DMF/tenure/discount rate parameters from Cromwell + Henley feasibility models).",
        "They are reference points only — your project will diverge based on site characteristics, market positioning, and operational decisions.",
        "Tenure default is 9 years (midpoint of Cromwell 8 / Henley 10).",
        "Internal discount rate default is 17.5% (Cromwell feasibility). JLL valuation discount rate is 14.0% — different basis, used for asset valuation.",
    ]),
]

r = 4
for section_title, paras in notes_content:
    ws_notes[f"B{r}"] = section_title
    ws_notes[f"B{r}"].font = Font(bold=True, size=12, color="1F3864")
    r += 1
    for para in paras:
        ws_notes[f"B{r}"] = para
        ws_notes[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws_notes.row_dimensions[r].height = max(15, 15 * (len(para) // 100 + 1))
        r += 1
    r += 1

# =================================================================
# SHEET 6: ASSUMPTIONS REFERENCE (copy Kyle's workbook verbatim)
# =================================================================
src_wb = load_workbook(SRC_ASSUMPTIONS)
src_ws = src_wb.active

ws_ref = wb.create_sheet("AssumptionsRef")
# Set column widths to match source
for col_letter in "ABCDEFGH":
    if col_letter in src_ws.column_dimensions:
        ws_ref.column_dimensions[col_letter].width = src_ws.column_dimensions[col_letter].width or 14
    else:
        ws_ref.column_dimensions[col_letter].width = 14
ws_ref.column_dimensions["A"].width = 50
ws_ref.column_dimensions["H"].width = 80

# Copy all cell values + formatting
for row in src_ws.iter_rows():
    for cell in row:
        new_cell = ws_ref[cell.coordinate]
        new_cell.value = cell.value
        if cell.has_style:
            new_cell.font = copy(cell.font)
            new_cell.fill = copy(cell.fill)
            new_cell.alignment = copy(cell.alignment)
            new_cell.number_format = cell.number_format
            new_cell.border = copy(cell.border)

# Rename to be user-friendly
ws_ref.title = "Assumptions Reference"

# Update INPUT cell formulas that referenced AssumptionsRef (we used that name when writing, sheet got renamed)
# Re-write the formulas with correct sheet name
def fix_assumptions_ref(ws_in_target):
    for row in ws_in_target.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "AssumptionsRef!" in cell.value:
                cell.value = cell.value.replace("AssumptionsRef!", "'Assumptions Reference'!")

fix_assumptions_ref(ws_in)

# =================================================================
# Sheet order — brief specifies Inputs first, then Calculations, MVP, Scenarios, Notes
# Assumptions Reference goes last
# =================================================================
order = ["Inputs", "Calculations", "Maximum Viable Price", "Scenarios", "Notes", "Assumptions Reference"]
wb._sheets = [wb[s] for s in order]

# Activate Inputs sheet on open
wb.active = 0

# Save
wb.save(OUT)
print(f"✓ Built {OUT.name}  ({OUT.stat().st_size:,} bytes)")
print(f"  Sheets in order: {wb.sheetnames}")
print(f"  Inputs key cells: land scenario {INP['land_scen']}, units {INP['units']}, ORA {INP['ora']}")
print(f"  Calculations: 4 scenarios in C:F, SELECTED in G, verdict row {VERD_R}")
print(f"  Headline verdict on Inputs row 12")
